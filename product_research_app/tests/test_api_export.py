from __future__ import annotations

from io import BytesIO
from pathlib import Path
import sys

import pytest
from openpyxl import load_workbook
from PIL import Image as PILImage

sys.path.append(str(Path(__file__).resolve().parents[2]))

from product_research_app import database, db
from product_research_app.api import app as flask_app
from product_research_app.api import export as export_api


import product_research_app.api as api_module


def _init_db(tmp_path: Path, monkeypatch):
    original_get_db = db.get_db
    target_path = str(tmp_path / "data.sqlite3")

    def _patched_get_db(path: str = "product_research_app/data.sqlite3", write: bool = False):
        actual = path
        if not path or path == "product_research_app/data.sqlite3":
            actual = target_path
        return original_get_db(actual, write=write)

    monkeypatch.setattr(db, "get_db", _patched_get_db)
    monkeypatch.setattr(api_module, "get_db", _patched_get_db)

    db.close_db()
    conn = db.get_db(target_path)
    database.initialize_database(conn)
    return conn


def test_export_generates_xlsx_with_images(tmp_path, monkeypatch):
    conn = _init_db(tmp_path, monkeypatch)
    product_id = database.insert_product(
        conn,
        name="Prod A",
        description="Description",
        category="Category",
        price=12.34,
        currency="USD",
        image_url="http://example.com/prod-a.png",
        source="Dropispy",
        desire="Superb",
        desire_magnitude="High",
        awareness_level="Solution-Aware",
        competition_level="Low",
        date_range="2024-01-01~2024-02-01",
        extra={
            "rating": 4.7,
            "units_sold": "1.2K",
            "revenue": "$500",
            "conversion_rate": "12%",
            "launch_date": "2024-01-01",
            "product_url": "https://shop.com/prod-a",
        },
    )
    conn.execute(
        "UPDATE products SET winner_score=?, winner_score_raw=? WHERE id=?",
        (88, 88.5, product_id),
    )
    conn.commit()
    assert getattr(db, "_DB_PATH", None) == str(tmp_path / "data.sqlite3")
    count = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    assert count == 1
    conn_again = db.get_db()
    assert conn_again is conn
    probe_rows = flask_app.config["ROW_PROVIDER"]([product_id], [])
    assert probe_rows and probe_rows[0]["id"] == product_id

    def fake_download(url: str, out_path: Path) -> bool:
        img = PILImage.new("RGBA", (export_api.IMG_W, export_api.IMG_W), (255, 0, 0, 255))
        img.save(out_path, format="PNG")
        return True

    monkeypatch.setattr(export_api, "_download_png_resized", fake_download)

    flask_app.config["TESTING"] = True
    with flask_app.test_client() as client:
        response = client.post(
            "/api/export",
            json={"ids": [product_id], "columns": [{"key": "name", "title": "Name"}]},
        )

    assert response.status_code == 200
    ctype = response.headers.get("Content-Type", "")
    assert ctype.startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disposition = response.headers.get("Content-Disposition", "")
    assert disposition.endswith(".xlsx") or ".xlsx" in disposition

    workbook = load_workbook(BytesIO(response.data))
    assert workbook.sheetnames == ["VEGA_INPUT", "PRODUCTS_FULL"]

    ws_in = workbook["VEGA_INPUT"]
    ws_full = workbook["PRODUCTS_FULL"]

    assert ws_in.freeze_panes == "A2"
    assert ws_full.freeze_panes == "A2"
    assert ws_in.max_row == 2
    assert ws_in["A2"].value == "Prod A"
    assert ws_in["B2"].value == "https://shop.com/prod-a"
    assert ws_in["C2"].value == "Superb"
    assert ws_in["D2"].value == "High"
    assert ws_in["E2"].value == "Solution-Aware"
    assert ws_in["F2"].value == "Low"

    assert ws_full["A2"].value == product_id
    assert ws_full["C2"].value == "Prod A"
    assert ws_full["D2"].value == "Category"
    assert ws_full["E2"].value == pytest.approx(12.34, rel=1e-6)
    assert ws_full["F2"].value == pytest.approx(4.7, rel=1e-6)
    assert ws_full["G2"].value == pytest.approx(1200, rel=1e-6)
    assert ws_full["H2"].value == pytest.approx(500, rel=1e-6)
    assert ws_full["I2"].value == pytest.approx(12, rel=1e-6)
    assert ws_full["J2"].value == "2024-01-01"
    assert ws_full["K2"].value == "2024-01-01~2024-02-01"
    assert ws_full["L2"].value == "Superb"
    assert ws_full["M2"].value == "High"
    assert ws_full["N2"].value == "Solution-Aware"
    assert ws_full["O2"].value == "Low"
    assert ws_full["P2"].value == pytest.approx(88, rel=1e-6)
    assert ws_full["Q2"].value == "https://shop.com/prod-a"

    assert ws_full.row_dimensions[2].height == pytest.approx(
        export_api.IMG_W + export_api.ROW_PAD
    )
    assert ws_full.column_dimensions["B"].width == pytest.approx(
        export_api._px_to_width(export_api.IMG_W + 20)
    )
    assert ws_full._images, "Product image should be embedded in the worksheet"

    workbook.close()
    db.close_db()
