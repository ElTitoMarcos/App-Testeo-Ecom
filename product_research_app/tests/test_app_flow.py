import io
import json
import logging
import sqlite3
import sys
import time
from pathlib import Path
from typing import List
from urllib.parse import urlparse

sys.path.append(str(Path(__file__).resolve().parents[2]))

from product_research_app import web_app, database, config, routes_export_minimal
from product_research_app.routes_export_minimal import COLUMNS as EXPORT_COLUMNS
from product_research_app.services import winner_score
from product_research_app.services import config as cfg_service
from product_research_app.utils.db import row_to_dict
from openpyxl.utils import get_column_letter

def setup_env(tmp_path, monkeypatch):
    monkeypatch.setattr(web_app, "DB_PATH", tmp_path / "data.sqlite3")
    monkeypatch.setattr(web_app, "LOG_DIR", tmp_path / "logs")
    monkeypatch.setattr(web_app, "LOG_PATH", tmp_path / "logs" / "app.log")
    web_app.LOG_DIR.mkdir(exist_ok=True)
    for h in list(logging.getLogger().handlers):
        logging.getLogger().removeHandler(h)
    logging.basicConfig(
        level=logging.INFO,
        handlers=[logging.FileHandler(web_app.LOG_PATH, encoding="utf-8")],
        force=True,
    )
    monkeypatch.setattr(config, "CONFIG_FILE", tmp_path / "config.json")
    monkeypatch.setattr(cfg_service, "DB_PATH", tmp_path / "data.sqlite3")
    cfg_service.init_app_config()
    return web_app.ensure_db()

def make_xlsx(path: Path, rows: List[List[object]]):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append([
        "name",
        "price",
        "rating",
        "units_sold",
        "revenue",
        "date_range",
        "image_count",
        "shipping_days",
        "profit_margin",
        "desire",
        "competition",
    ])
    for r in rows:
        ws.append(r)
    wb.save(path)

def test_app_startup(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    assert web_app.LOG_PATH.exists()
    cur = conn.cursor()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='scores'")
    assert cur.fetchone() is not None

def test_import_generates_scores(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(config, "is_auto_fill_ia_on_import_enabled", lambda: False)
    xlsx = tmp_path / "products.xlsx"
    make_xlsx(
        xlsx,
        [
            ["Prod1", 10, 4.5, 100, 1000, "2024-01-01~2024-02-01", 3, 5, 0.3, "High", "Low"],
            ["Prod2", 20, 3.0, 50, 500, "2024-03-01~2024-04-01", 2, 7, 0.2, "Medium", "Medium"],
        ],
    )
    job_id = database.create_import_job(conn, str(xlsx))
    web_app._process_import_job(job_id, xlsx, "products.xlsx")
    deadline = time.time() + 5
    while time.time() < deadline:
        job_row = database.get_import_job(conn, job_id)
        payload = web_app._job_payload_from_row(job_row)
        if payload and payload.get("phase") == "done" and payload.get("status") == "done":
            break
        time.sleep(0.05)
    else:
        raise AssertionError("import job did not complete in time")
    products = [row_to_dict(r) for r in database.list_products(conn)]
    assert len(products) == 2
    for p in products:
        assert 0 <= p.get("winner_score", 0) <= 100

def test_scoring_v2_generate_cases(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(web_app, "ensure_db", lambda: conn)
    pid_a = database.insert_product(
        conn,
        name="A",
        description="",
        category="",
        price=None,
        currency=None,
        image_url="",
        source="",
        extra={},
        product_id=1,
    )
    pid_b = database.insert_product(
        conn,
        name="B",
        description="",
        category="",
        price=20.0,
        currency=None,
        image_url="",
        source="",
        extra={"rating": 4.0},
        product_id=2,
    )
    body = json.dumps({"ids": [pid_a, pid_b]})
    class Dummy:
        def __init__(self, body):
            self.headers = {"Content-Length": str(len(body))}
            self.rfile = io.BytesIO(body.encode("utf-8"))
            self.wfile = io.BytesIO()
            self.path = "/api/winner-score/generate"
        def _set_json(self, code=200):
            self.status = code

    handler = Dummy(body)
    web_app.RequestHandler.handle_scoring_v2_generate(handler)
    resp = json.loads(handler.wfile.getvalue().decode("utf-8"))
    assert resp["ok"] is True
    assert resp["processed"] == 2
    assert resp["updated"] == 2
    assert resp["weights_all"]
    assert resp["weights_eff"]
    prod_a = database.get_product(conn, pid_a)
    prod_b = database.get_product(conn, pid_b)
    assert 0 <= prod_a["winner_score"] <= 100
    assert 0 <= prod_b["winner_score"] <= 100

    body2 = json.dumps({"ids": [pid_b]})
    handler2 = Dummy(body2)
    web_app.RequestHandler.handle_scoring_v2_generate(handler2)
    resp2 = json.loads(handler2.wfile.getvalue().decode("utf-8"))
    assert resp2["ok"] is True
    assert resp2["processed"] == 1
    assert resp2["updated"] == 0
    assert resp2["weights_all"]
    assert resp2["weights_eff"]


def test_scoring_v2_generate_all_when_no_ids(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(web_app, "ensure_db", lambda: conn)
    pid_a = database.insert_product(
        conn,
        name="A",
        description="",
        category="",
        price=None,
        currency=None,
        image_url="",
        source="",
        extra={"rating": 4.0},
        product_id=1,
    )
    pid_b = database.insert_product(
        conn,
        name="B",
        description="",
        category="",
        price=None,
        currency=None,
        image_url="",
        source="",
        extra={},
        product_id=2,
    )

    body = json.dumps({})

    class Dummy:
        def __init__(self, body):
            self.headers = {"Content-Length": str(len(body))}
            self.rfile = io.BytesIO(body.encode("utf-8"))
            self.wfile = io.BytesIO()
            self.path = "/api/winner-score/generate"

        def _set_json(self, code=200):
            self.status = code

    handler = Dummy(body)
    web_app.RequestHandler.handle_scoring_v2_generate(handler)
    resp = json.loads(handler.wfile.getvalue().decode("utf-8"))
    assert resp["processed"] == 2
    assert resp["weights_all"]
    assert resp["weights_eff"]

def test_products_endpoint_serializes_rows(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(web_app, "ensure_db", lambda: conn)
    pid = database.insert_product(
        conn,
        name="X",
        description="",
        category="",
        price=0.0,
        currency=None,
        image_url="",
        source="",
        extra={},
        product_id=1,
    )
    conn.execute("UPDATE products SET winner_score=42 WHERE id=?", (pid,))
    conn.commit()

    class Dummy:
        def __init__(self):
            self.path = "/products"
            self.headers = {}
            self.wfile = io.BytesIO()

        def _set_json(self, code=200):
            self.status = code

        def send_error(self, code, msg=None):
            raise AssertionError(f"error {code}")

        def safe_write(self, func):
            try:
                func()
                return True
            except Exception:
                return False

    handler = Dummy()
    web_app.RequestHandler.do_GET(handler)
    resp = json.loads(handler.wfile.getvalue().decode("utf-8"))
    assert isinstance(resp, list) and resp and resp[0]["id"] == pid
    assert resp[0]["winner_score"] == 42


def test_desire_serialization_and_logging(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(web_app, "ensure_db", lambda: conn)
    pid1 = database.insert_product(
        conn,
        name="P1",
        description="",
        category="",
        price=1.0,
        currency=None,
        image_url="",
        source="",
        desire="Top",
        extra={},
        product_id=1,
    )
    pid2 = database.insert_product(
        conn,
        name="P2",
        description="",
        category="",
        price=2.0,
        currency=None,
        image_url="",
        source="",
        extra={"desire": "Extra"},
        product_id=2,
    )
    pid3 = database.insert_product(
        conn,
        name="P3",
        description="",
        category="",
        price=3.0,
        currency=None,
        image_url="",
        source="",
        extra={},
        product_id=3,
    )

    class Dummy:
        def __init__(self, path):
            self.path = path
            self.headers = {}
            self.wfile = io.BytesIO()

        def _set_json(self, code=200):
            self.status = code

        def send_error(self, code, msg=None):
            raise AssertionError(f"error {code}")

        def safe_write(self, func):
            try:
                func()
                return True
            except Exception:
                return False

    handler = Dummy("/products")
    web_app.RequestHandler.do_GET(handler)
    resp = json.loads(handler.wfile.getvalue().decode("utf-8"))
    p1 = next(p for p in resp if p["id"] == pid1)
    p2 = next(p for p in resp if p["id"] == pid2)
    p3 = next(p for p in resp if p["id"] == pid3)
    assert p1["desire"] == "Top"
    assert isinstance(p1["price"], (int, float))
    assert p2["desire"] == "Extra"
    assert p2["extras"].get("desire") == "Extra"
    assert p3["desire"] is None
    log_text = web_app.LOG_PATH.read_text()
    assert f"desire_missing=true" in log_text and f"product={pid3}" in log_text

    lid = database.create_list(conn, "L")
    database.add_product_to_list(conn, lid, pid1)
    handler_list = Dummy(f"/list/{lid}")
    web_app.RequestHandler.do_GET(handler_list)
    resp_list = json.loads(handler_list.wfile.getvalue().decode("utf-8"))
    assert resp_list and resp_list[0]["desire"] == "Top"


def test_patch_product_desire(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(web_app, "ensure_db", lambda: conn)
    pid = database.insert_product(
        conn,
        name="P1",
        description="",
        category="",
        price=1.0,
        currency=None,
        image_url="",
        source="",
        desire="old",
    )
    body = json.dumps({"desire": "nuevo texto", "unknown": "x"})

    class Dummy:
        def __init__(self, path, body):
            self.path = path
            self.headers = {"Content-Length": str(len(body))}
            self.rfile = io.BytesIO(body.encode("utf-8"))
            self.wfile = io.BytesIO()

        def _set_json(self, code=200):
            self.status = code

    handler = Dummy(f"/api/products/{pid}", body)
    web_app.RequestHandler.do_PATCH(handler)
    resp = json.loads(handler.wfile.getvalue().decode("utf-8"))
    assert handler.status == 200
    assert resp["desire"] == "nuevo texto"
    assert database.get_product(conn, pid)["desire"] == "nuevo texto"

    handler_nf = Dummy("/api/products/9999", body)
    web_app.RequestHandler.do_PATCH(handler_nf)
    assert handler_nf.status == 404
    log_text = web_app.LOG_PATH.read_text()
    assert "PATCH not found product_id=9999" in log_text

def test_patch_winner_weights_persists(tmp_path, monkeypatch):
    setup_env(tmp_path, monkeypatch)

    body = json.dumps({"weights": {"rating": 25}, "order": ["rating", "price"]})

    class Dummy:
        def __init__(self, body):
            self.path = "/api/config/winner-weights"
            self.headers = {"Content-Length": str(len(body))}
            self.rfile = io.BytesIO(body.encode("utf-8"))
            self.wfile = io.BytesIO()

        def _set_json(self, code=200):
            self.status = code

    handler = Dummy(body)
    web_app.RequestHandler.do_PATCH(handler)
    resp = json.loads(handler.wfile.getvalue().decode("utf-8"))
    assert resp["weights"]["rating"] == 25
    assert resp["order"][0] == "rating"
    from product_research_app.services.config import get_winner_weights_raw, get_winner_order_raw
    data = get_winner_weights_raw()
    order = get_winner_order_raw()
    assert data.get("rating") == 25
    assert order[0] == "rating"

def test_config_oldness_preference_roundtrip(tmp_path, monkeypatch):
    setup_env(tmp_path, monkeypatch)

    class DummyGet:
        def __init__(self):
            self.path = "/config"
            self.headers = {}
            self.rfile = io.BytesIO()
            self.wfile = io.BytesIO()

        def _set_json(self, code=200):
            self.status = code

    g = DummyGet()
    web_app.RequestHandler.do_GET(g)
    resp = json.loads(g.wfile.getvalue().decode("utf-8"))
    assert resp.get("oldness_preference") == "newer"

    body = json.dumps({"oldness_preference": "older"})

    class DummyPost:
        def __init__(self, body):
            self.path = "/setconfig"
            self.headers = {"Content-Length": str(len(body))}
            self.rfile = io.BytesIO(body.encode("utf-8"))
            self.wfile = io.BytesIO()

        def _set_json(self, code=200):
            self.status = code

    p = DummyPost(body)
    web_app.RequestHandler.handle_setconfig(p)
    assert p.status == 200
    assert config.load_config().get("oldness_preference") == "older"

def test_get_endpoints_return_json(tmp_path, monkeypatch):
    setup_env(tmp_path, monkeypatch)
    from http.server import HTTPServer
    import threading, urllib.request, json, time

    server = HTTPServer(("127.0.0.1", 0), web_app.RequestHandler)
    port = server.server_address[1]
    thread = threading.Thread(target=server.serve_forever)
    thread.daemon = True
    thread.start()
    try:
        time.sleep(0.1)
        resp = urllib.request.urlopen(f"http://127.0.0.1:{port}/config")
        assert resp.status == 200
        cfg = json.loads(resp.read().decode("utf-8"))
        weights = cfg.get("weights", {})
        assert all(0 <= v <= 100 for v in weights.values())
        assert set(weights.keys()) == set(config.SCORING_DEFAULT_WEIGHTS.keys())
        resp = urllib.request.urlopen(f"http://127.0.0.1:{port}/lists")
        assert resp.status == 200
        json.loads(resp.read().decode("utf-8"))
    finally:
        server.shutdown()
        thread.join()


def test_weight_changes_persist_without_score_reset(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(web_app, "ensure_db", lambda: conn)
    pid = database.insert_product(
        conn,
        name="A",
        description="",
        category="",
        price=None,
        currency=None,
        image_url="",
        source="",
        extra={},
        product_id=1,
    )
    conn.execute("UPDATE products SET winner_score=55 WHERE id=?", (pid,))
    conn.commit()

    def patch_weight(key, value):
        body = json.dumps({"weights": {key: value}})
        class Dummy:
            def __init__(self, body):
                self.path = "/api/config/winner-weights"
                self.headers = {"Content-Length": str(len(body))}
                self.rfile = io.BytesIO(body.encode("utf-8"))
                self.wfile = io.BytesIO()

            def _set_json(self, code=200):
                self.status = code

        h = Dummy(body)
        web_app.RequestHandler.do_PATCH(h)
        resp = json.loads(h.wfile.getvalue().decode("utf-8"))
        assert resp["weights"][key] == value

    patch_weight("rating", 20)
    patch_weight("price", 30)
    patch_weight("units_sold", 50)
    from product_research_app.services.config import get_winner_weights_raw

    weights = get_winner_weights_raw()
    assert weights.get("rating") == 20
    assert weights.get("price") == 30
    assert weights.get("units_sold") == 50
    prod = database.get_product(conn, pid)
    assert prod["winner_score"] == 55


def test_import_new_batch_preserves_existing_scores(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(web_app, "ensure_db", lambda: conn)
    monkeypatch.setattr(config, "is_auto_fill_ia_on_import_enabled", lambda: False)
    pid = database.insert_product(
        conn,
        name="Old", description="", category="", price=None, currency=None,
        image_url="", source="", extra={}, product_id=1,
    )
    conn.execute("UPDATE products SET winner_score=10 WHERE id=?", (pid,))
    conn.commit()

    xlsx = tmp_path / "batch.xlsx"
    make_xlsx(xlsx, [["New", 10, 4.5, 100, 1000, 50, 3, 5, 0.3, "High", "Low"]])
    job_id = database.create_import_job(conn, str(xlsx))
    web_app._process_import_job(job_id, xlsx, "batch.xlsx")
    deadline = time.time() + 5
    while time.time() < deadline:
        job_row = database.get_import_job(conn, job_id)
        payload = web_app._job_payload_from_row(job_row)
        if payload and payload.get("phase") == "done" and payload.get("status") == "done":
            break
        time.sleep(0.05)
    else:
        raise AssertionError("import job did not finish")

    prod_old = database.get_product(conn, pid)
    assert prod_old["winner_score"] == 10
    cur = conn.execute("SELECT id, winner_score FROM products WHERE id != ?", (pid,))
    rows = cur.fetchall()
    assert rows
    for _id, score in rows:
        assert score is not None and score > 0


def test_recalculate_selected_and_all(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(web_app, "ensure_db", lambda: conn)
    ids = []
    for i in range(8):
        pid = database.insert_product(
            conn,
            name=f"P{i}", description="", category="", price=None, currency=None,
            image_url="", source="", extra={"rating": 4.5}, product_id=i + 1,
        )
        conn.execute("UPDATE products SET winner_score=1 WHERE id=?", (pid,))
        ids.append(pid)
    conn.commit()

    subset = ids[:5]
    body = json.dumps({"product_ids": subset})
    class Dummy:
        def __init__(self, body):
            self.headers = {"Content-Length": str(len(body))}
            self.rfile = io.BytesIO(body.encode("utf-8"))
            self.wfile = io.BytesIO()
            self.path = "/api/winner-score/generate"

        def _set_json(self, code=200):
            self.status = code

    h = Dummy(body)
    web_app.RequestHandler.handle_scoring_v2_generate(h)
    json.loads(h.wfile.getvalue().decode("utf-8"))
    for pid in subset:
        assert database.get_product(conn, pid)["winner_score"] != 1
    for pid in ids[5:]:
        assert database.get_product(conn, pid)["winner_score"] == 1

    body2 = json.dumps({})
    h2 = Dummy(body2)
    web_app.RequestHandler.handle_scoring_v2_generate(h2)
    json.loads(h2.wfile.getvalue().decode("utf-8"))
    for pid in ids:
        assert database.get_product(conn, pid)["winner_score"] != 1


def test_weights_hash_changes_after_patch(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(web_app, "ensure_db", lambda: conn)

    pid = database.insert_product(
        conn,
        name="P0",
        description="",
        category="",
        price=10.0,
        currency=None,
        image_url="",
        source="",
        extra={"rating": 4.5},
        product_id=1,
    )
    conn.execute("UPDATE products SET winner_score=1 WHERE id=?", (pid,))
    conn.commit()

    body = json.dumps({})

    class Dummy:
        def __init__(self, body):
            self.headers = {"Content-Length": str(len(body))}
            self.rfile = io.BytesIO(body.encode("utf-8"))
            self.wfile = io.BytesIO()
            self.path = "/api/winner-score/generate"

        def _set_json(self, code=200):
            self.status = code

    h1 = Dummy(body)
    web_app.RequestHandler.handle_scoring_v2_generate(h1)
    resp1 = json.loads(h1.wfile.getvalue().decode("utf-8"))
    hash1 = resp1["weights_all"]
    eff1 = resp1["weights_eff"]
    ver1 = config.get_weights_version()

    body_patch = json.dumps({"weights": {"rating": 20}})

    class Patcher:
        def __init__(self, body):
            self.path = "/api/config/winner-weights"
            self.headers = {"Content-Length": str(len(body))}
            self.rfile = io.BytesIO(body.encode("utf-8"))
            self.wfile = io.BytesIO()

        def _set_json(self, code=200):
            self.status = code

    p = Patcher(body_patch)
    web_app.RequestHandler.do_PATCH(p)

    h2 = Dummy(body)
    web_app.RequestHandler.handle_scoring_v2_generate(h2)
    resp2 = json.loads(h2.wfile.getvalue().decode("utf-8"))
    assert resp2["weights_all"] != hash1
    assert resp2["weights_eff"] != eff1
    assert config.get_weights_version() > ver1

def test_logging_and_explain_endpoint(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(web_app, "ensure_db", lambda: conn)
    web_app.DEBUG = True

    config.update_weight("oldness", 2.0)

    pid = database.insert_product(
        conn,
        name="L",
        description="",
        category="",
        price=None,
        currency=None,
        image_url="",
        source="",
        extra={"rating": 4.0},
        product_id=1,
    )

    body = json.dumps({"ids": [pid]})

    class DummyGen:
        def __init__(self, body):
            self.headers = {"Content-Length": str(len(body))}
            self.rfile = io.BytesIO(body.encode("utf-8"))
            self.wfile = io.BytesIO()
            self.path = "/api/winner-score/generate"

        def _set_json(self, code=200):
            self.status = code

    handler = DummyGen(body)
    web_app.RequestHandler.handle_scoring_v2_generate(handler)

    log_text = web_app.LOG_PATH.read_text()
    assert "oldness" in log_text
    assert "effective_weights={" in log_text
    assert "order=" in log_text

    parsed = urlparse(f"/api/winner-score/explain?ids={pid}")

    class DummyExplain:
        def __init__(self):
            self.wfile = io.BytesIO()

        def _set_json(self, code=200):
            self.status = code

    handler2 = DummyExplain()
    web_app.RequestHandler.handle_winner_score_explain(handler2, parsed)
    resp = json.loads(handler2.wfile.getvalue().decode("utf-8"))
    info = resp[str(pid)]
    assert "rating" in info["present"]
    assert "oldness" in info["missing"]
    eff = info["effective_weights"]
    assert set(eff.keys()) == set(winner_score.ALLOWED_FIELDS)
    assert abs(sum(eff.values()) - 1.0) < 1e-2


def test_weights_eff_stable_when_touching_missing_metric(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(web_app, "ensure_db", lambda: conn)

    pid = database.insert_product(
        conn,
        name="MM",
        description="",
        category="",
        price=None,
        currency=None,
        image_url="",
        source="",
        extra={"rating": 4.5},
        product_id=1,
    )
    conn.commit()

    config.update_weight("rating", 0.0)

    body = json.dumps({"product_ids": [pid]})

    class Dummy:
        def __init__(self, body, path="/api/winner-score/generate?debug=1"):
            self.path = path
            self.headers = {"Content-Length": str(len(body))}
            self.rfile = io.BytesIO(body.encode("utf-8"))
            self.wfile = io.BytesIO()

        def _set_json(self, code=200):
            self.status = code

    h1 = Dummy(body)
    web_app.RequestHandler.handle_scoring_v2_generate(h1)
    resp1 = json.loads(h1.wfile.getvalue().decode("utf-8"))
    hash_all1 = resp1["weights_all"]
    hash_eff1 = resp1["weights_eff"]
    sum1 = resp1["diag"]["sum_filtered"]
    assert sum1 > 0.0

    body_patch = json.dumps({"weights": {"units_sold": 20}})
    class Patcher:
        def __init__(self, body):
            self.path = "/api/config/winner-weights"
            self.headers = {"Content-Length": str(len(body))}
            self.rfile = io.BytesIO(body.encode("utf-8"))
            self.wfile = io.BytesIO()

        def _set_json(self, code=200):
            self.status = code

    p = Patcher(body_patch)
    web_app.RequestHandler.do_PATCH(p)

    h2 = Dummy(body)
    web_app.RequestHandler.handle_scoring_v2_generate(h2)
    resp2 = json.loads(h2.wfile.getvalue().decode("utf-8"))
    assert resp2["weights_all"] != hash_all1
    assert resp2["weights_eff"] != hash_eff1
    assert resp2["diag"]["sum_filtered"] > sum1


def _make_export_dummy(body: str):
    class Dummy:
        def __init__(self, payload: str):
            self.headers = {"Content-Length": str(len(payload))}
            self.rfile = io.BytesIO(payload.encode("utf-8"))
            self.wfile = io.BytesIO()
            self.sent_headers = {}
            self.status = None
            self.json_payload = None
            self.path = "/api/export/kalodata-minimal"

        def send_json(self, obj, status=200):
            self.status = status
            self.json_payload = obj

        def send_response(self, code):
            self.status = code

        def send_header(self, key, value):
            self.sent_headers[key] = value

        def end_headers(self):
            pass

    return Dummy(body)


def test_export_kalodata_minimal_requires_ids(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(web_app, "ensure_db", lambda: conn)

    handler = _make_export_dummy(json.dumps({}))
    web_app.RequestHandler.handle_export_kalodata_minimal(handler)

    assert handler.status == 400
    assert handler.json_payload == {"error": "ids_required"}


def test_export_kalodata_minimal_not_found(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(web_app, "ensure_db", lambda: conn)

    handler = _make_export_dummy(json.dumps({"ids": [999]}))
    web_app.RequestHandler.handle_export_kalodata_minimal(handler)

    assert handler.status == 404
    assert handler.json_payload == {"error": "products_not_found"}


def test_export_kalodata_minimal_success(tmp_path, monkeypatch):
    conn = setup_env(tmp_path, monkeypatch)
    monkeypatch.setattr(web_app, "ensure_db", lambda: conn)
    monkeypatch.setattr(routes_export_minimal, "STATE_DIR", tmp_path / "state")
    monkeypatch.setattr(routes_export_minimal, "EXPORT_DIR", tmp_path / "exports")

    extra1 = {
        "TikTokUrl": "https://tiktok.example/alpha",
        "KalodataUrl": "https://kalodata.example/alpha",
        "Img_url": "https://img.example/alpha.jpg",
        "Category": "Beauty",
        "Price($)": "$19.99",
        "Product Rating": "4.8",
        "Item Sold": "1,200",
        "Revenue($)": "1000",
        "Live Revenue($)": "200",
        "Video Revenue($)": "300",
        "Shopping Mall Revenue($)": "50",
        "Launch Date": "2024-01-01",
        "desire_score": 0.8,
        "awareness_level": 2,
        "competition": 0.3,
    }
    extra2 = {
        "tiktok_url": "https://tiktok.example/beta",
        "KalodataUrl": "https://kalodata.example/beta",
        "image_url": "https://img.example/beta.png",
        "Category": "Home",
        "price": "29",
        "rating": 4.1,
        "units_sold": 80,
        "Revenue($)": "400",
        "Video Revenue($)": "100",
        "Launch Date": "2023-12-15",
        "desire_magnitude": 65,
        "customer_desire": "Resuelve dolor",
        "awareness_level_label": "Product-aware",
        "competition_level_label": "Medium",
    }
    extra3 = {
        "tiktok_url": "https://tiktok.example/gamma",
        "KalodataUrl": "https://kalodata.example/gamma",
        "Img_url": "https://img.example/gamma-missing.png",
        "Category": "Kitchen",
        "price": "15.50",
        "rating": 4.5,
        "units_sold": 500,
        "Revenue($)": "250",
        "Live Revenue($)": "30",
        "Video Revenue($)": "20",
        "Launch Date": "2023-11-30",
        "desires": "Resolver frustración",
        "awareness_level": 4,
        "competition_level": 80,
    }

    pid1 = database.insert_product(
        conn,
        name="Alpha",
        description="",
        category="Beauty",
        price=19.99,
        currency="USD",
        image_url="https://img.example/alpha.jpg",
        source="kalodata",
        desire="Fuerte deseo",
        desire_magnitude=None,
        awareness_level=None,
        competition_level=None,
        extra=extra1,
    )
    pid2 = database.insert_product(
        conn,
        name="Beta",
        description="",
        category="Home",
        price=None,
        currency=None,
        image_url="https://img.example/beta.png",
        source="kalodata",
        desire=None,
        desire_magnitude=None,
        awareness_level=None,
        competition_level=None,
        extra=extra2,
    )
    pid3 = database.insert_product(
        conn,
        name="Gamma",
        description="",
        category="Kitchen",
        price=None,
        currency=None,
        image_url="https://img.example/gamma-missing.png",
        source="kalodata",
        desire=None,
        desire_magnitude=None,
        awareness_level=None,
        competition_level=None,
        extra=extra3,
    )

    routes_export_minimal._IMG_CACHE.clear()

    from PIL import Image as PILImage

    def fake_fetch(url):
        if url and "missing" not in str(url):
            img = PILImage.new("RGB", (400, 260), color=(255, 0, 0))
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            buf.seek(0)
            return buf
        return None

    monkeypatch.setattr(routes_export_minimal, "_fetch_and_resize", fake_fetch)

    handler = _make_export_dummy(json.dumps({"ids": [pid1, pid2, pid3]}))
    web_app.RequestHandler.handle_export_kalodata_minimal(handler)

    assert handler.status == 200
    disposition = handler.sent_headers.get("Content-Disposition", "")
    assert disposition == "attachment; filename=Analisis_Export_0001.xlsx"

    payload = handler.wfile.getvalue()
    assert payload

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(payload))
    assert wb.sheetnames == ["products"]
    ws = wb["products"]
    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    assert header == list(EXPORT_COLUMNS)

    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == ws.dimensions

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 3
    row1, row2, row3 = rows

    assert row1[0] is None
    assert row2[0] is None
    assert row3[0] == "(no image)"
    assert row1[1] == "Alpha"
    assert row2[1] == "Beta"
    assert row3[1] == "Gamma"
    assert row1[2] == "Beauty"
    assert row2[2] == "Home"
    assert row3[2] == "Kitchen"
    assert row1[3] and row1[3].strftime("%Y-%m-%d") == "2024-01-01"
    assert row2[3] and row2[3].strftime("%Y-%m-%d") == "2023-12-15"
    assert row3[3] and row3[3].strftime("%Y-%m-%d") == "2023-11-30"
    assert row1[4] == 19.99
    assert row2[4] == 29.0
    assert row3[4] == 15.5
    assert row1[5] == 4.8
    assert row2[5] == 4.1
    assert row3[5] == 4.5
    assert row1[6] == 1550.0
    assert row2[6] == 500.0
    assert row3[6] == 300.0
    assert row1[7] == 1200
    assert row2[7] == 80
    assert row3[7] == 500
    assert row1[8] == "Fuerte deseo"
    assert row2[8] == "Resuelve dolor"
    assert row3[8] == "Resolver frustración"
    assert row1[9] == 80
    assert row2[9] == 65
    assert row3[9] == 50
    assert row1[10] == "Solution-aware"
    assert row2[10] == "Product-aware"
    assert row3[10] == "Most aware"
    assert row1[11] == "Low"
    assert row2[11] == "Medium"
    assert row3[11] == "High"
    assert row1[12] == "https://tiktok.example/alpha"
    assert row2[12] == "https://tiktok.example/beta"
    assert row3[12] == "https://tiktok.example/gamma"
    assert row1[13] == "https://kalodata.example/alpha"
    assert row2[13] == "https://kalodata.example/beta"
    assert row3[13] == "https://kalodata.example/gamma"
    assert row1[14] == "https://img.example/alpha.jpg"
    assert row2[14] == "https://img.example/beta.png"
    assert row3[14] == "https://img.example/gamma-missing.png"

    assert ws.column_dimensions["A"].width == 38
    assert ws.column_dimensions["I"].width == 45
    assert ws.column_dimensions["O"].width == 40

    assert "tbl_products" in ws.tables
    assert ws.tables["tbl_products"].ref == "A1:O4"

    dv_ranges = [dv.sqref for dv in ws.data_validations.dataValidation]
    assert any(str(dv) == "J2:J4" for dv in dv_ranges)

    cf_rules = [
        rule
        for cf in ws.conditional_formatting
        for rule in cf.rules
    ]
    assert not any(getattr(rule, "type", "") == "dataBar" for rule in cf_rules)

    dm_index = header.index("desire magnitude")
    dm_letter = get_column_letter(dm_index + 1)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws[f"{dm_letter}{row_idx}"]
        assert cell.fill.fill_type is None

    saved_files = sorted((routes_export_minimal.EXPORT_DIR).glob("Analisis_Export_*.xlsx"))
    assert saved_files
    assert saved_files[0].name == "Analisis_Export_0001.xlsx"
    seq_path = routes_export_minimal.STATE_DIR / "export_seq.json"
    assert seq_path.exists()
    with seq_path.open("r", encoding="utf-8") as fh:
        seq_payload = json.load(fh)
    assert seq_payload == {"export_seq": 1}

    assert len(getattr(ws, "_images", [])) == 2
    assert ws.row_dimensions[2].height == routes_export_minimal._IMG_ROW_HEIGHT
    assert ws.row_dimensions[4].height == 45
