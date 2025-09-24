"""
Simple web interface for the Product Research Copilot.

This module starts an HTTP server that serves a minimal single‑page application
allowing the user to upload CSV/JSON files, configure OpenAI settings, trigger
batch evaluations, and make custom GPT queries.  The UI is written in
plain HTML/JavaScript and features a dark mode.  It is intended to be
platform‑agnostic and requires only the Python standard library.

Limitations:
    - OCR on image uploads is not implemented; when uploading images the user
      will need to input product details manually in the app.
    - For large datasets the evaluation may block the server; consider running
      the batch importer beforehand.

Usage:
    python -m product_research_app.web_app [--host 127.0.0.1] [--port 8000]
Then open http://host:port in a browser.
"""

from __future__ import annotations

import json
import os
import io
import re
import logging
import requests
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
from pathlib import Path
from email.parser import BytesParser
from email.policy import default
import threading
import time
import sqlite3
import math
import hashlib
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Union

from . import database
from .db import get_db, get_last_performance_config
from . import config
from .services import ai_columns
from .services import winner_score as winner_calc
from .services import trends_service
from .services.config import get_default_winner_weights
from .services.importer_fast import DEFAULT_BATCH_SIZE, fast_import, fast_import_records
from . import gpt
from .prompts.registry import normalize_task
from . import title_analyzer
from . import product_enrichment
from .sse import publish_progress
from .utils import sanitize_product_name
from .utils.db import row_to_dict, rget

WINNER_SCORE_FIELDS = list(winner_calc.FEATURE_MAP.keys())

APP_DIR = Path(__file__).resolve().parent
DB_PATH = APP_DIR / "data.sqlite3"
STATIC_DIR = APP_DIR / "static"
ROOT_DIR = APP_DIR.parent
LOG_DIR = ROOT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)
LOG_PATH = LOG_DIR / "app.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

DEBUG = bool(os.environ.get("DEBUG"))

DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y")


DEFAULT_ORDER_LIST = list(config.DEFAULT_WINNER_ORDER)


def _clamp_weight_value(value: Any) -> int:
    try:
        return int(max(0, min(100, float(value))))
    except Exception:
        return 0


def _sanitize_weights_map(raw: Optional[Dict[str, Any]]) -> Dict[str, int]:
    sanitized: Dict[str, int] = {}
    base = raw or {}
    for key in DEFAULT_ORDER_LIST:
        sanitized[key] = _clamp_weight_value(base.get(key, 50))
    for key, value in base.items():
        if key not in sanitized:
            sanitized[key] = _clamp_weight_value(value)
    return sanitized


def _normalize_order_list(order, available: Dict[str, Any]) -> List[str]:
    seen: set[str] = set()
    out: List[str] = []
    keys = list(available.keys())
    for key in order or []:
        if key in available and key not in seen:
            out.append(key)
            seen.add(key)
    for key in DEFAULT_ORDER_LIST:
        if key in available and key not in seen:
            out.append(key)
            seen.add(key)
    for key in keys:
        if key not in seen:
            out.append(key)
            seen.add(key)
    return out

def _apply_weights_reset(existing: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    cfg = dict(existing or config.load_config())
    default_weights = get_default_winner_weights()
    cfg["winner_weights"] = dict(default_weights)
    order = list(config.DEFAULT_WINNER_ORDER)
    cfg["winner_order"] = order[:]
    cfg["weights_order"] = order[:]
    raw_enabled = cfg.get("weights_enabled")
    if isinstance(raw_enabled, dict):
        enabled = {k: bool(raw_enabled.get(k, True)) for k in default_weights.keys()}
    else:
        enabled = {k: True for k in default_weights.keys()}
    cfg["weights_enabled"] = enabled
    cfg["weightsUpdatedAt"] = int(time.time())
    config.save_config(cfg)
    try:
        winner_calc.invalidate_weights_cache()
    except Exception as exc:  # pragma: no cover - defensive logging
        logger.warning("reset invalidate failed: %s", exc)
    return cfg


def _build_weights_payload(cfg: Dict[str, Any]) -> Dict[str, Any]:
    payload = {
        "ok": True,
        "weights": dict(cfg.get("winner_weights", {})),
        "winner_weights": dict(cfg.get("winner_weights", {})),
        "order": list(cfg.get("winner_order", [])),
        "winner_order": list(cfg.get("winner_order", [])),
        "weights_order": list(cfg.get("weights_order", [])),
        "weights_enabled": dict(cfg.get("weights_enabled", {})),
    }
    if "weightsUpdatedAt" in cfg:
        payload["weightsUpdatedAt"] = cfg["weightsUpdatedAt"]
    if "weightsVersion" in cfg:
        payload["weightsVersion"] = cfg["weightsVersion"]
    return payload

def _sanitize_enabled_map(raw_enabled, keys: List[str]) -> Dict[str, bool]:
    if not isinstance(raw_enabled, dict):
        return {k: True for k in keys}
    return {k: bool(raw_enabled.get(k, True)) for k in keys}


_DB_INIT = False
_DB_INIT_PATH: Optional[str] = None
_DB_INIT_LOCK = threading.Lock()

IMPORT_STATUS: Dict[str, Dict[str, Any]] = {}
_IMPORT_STATUS_LOCK = threading.Lock()

_ENRICH_WORKERS: Dict[int, threading.Thread] = {}
_ENRICH_LOCK = threading.Lock()


def _parse_date(s: str):
    s = (s or "").strip()
    if not s:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def ensure_db():
    global _DB_INIT, _DB_INIT_PATH

    target_path = str(DB_PATH)
    conn = get_db(target_path)
    if not _DB_INIT or _DB_INIT_PATH != target_path:
        with _DB_INIT_LOCK:
            if not _DB_INIT or _DB_INIT_PATH != target_path:
                try:
                    database.initialize_database(conn)
                except Exception:
                    logger.exception("Database initialization failed")
                    raise
                try:
                    cur = conn.cursor()
                    cur.execute(
                        "DELETE FROM products WHERE id IN (1,2,3) OR lower(name) LIKE '%test%' OR lower(name) LIKE '%prueba%'"
                    )
                    conn.commit()
                except Exception:
                    pass
                logger.info("Database ready at %s", DB_PATH)
                _DB_INIT = True
                _DB_INIT_PATH = target_path
    return conn


def _update_import_status(task_id: str, **updates) -> Dict[str, Any]:
    with _IMPORT_STATUS_LOCK:
        state = IMPORT_STATUS.setdefault(task_id, {})
        state.update(updates)
        snapshot = dict(state)
    publish_progress({"event": "import", "task_id": task_id, **snapshot})
    return snapshot


def _set_import_progress(
    task_id: str,
    pct: Optional[Union[float, int]] = None,
    message: Optional[str] = None,
    **updates: Any,
) -> Dict[str, Any]:
    payload: Dict[str, Any] = {}
    if pct is not None:
        try:
            pct_val = int(round(float(pct)))
            payload["pct"] = max(0, min(100, pct_val))
        except Exception:
            pass
    if message is not None:
        payload["message"] = message
    payload.update(updates)
    return _update_import_status(task_id, **payload)


def _maybe_json(value):
    if value in (None, ""):
        return None
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except Exception:
        return None


def _job_payload_from_row(row):
    if row is None:
        return None
    data = row_to_dict(row)
    job_id = data.get("task_id") or data.get("id")
    if job_id is not None:
        try:
            job_int = int(job_id)
        except Exception:
            job_int = job_id
        data["job_id"] = job_int
        data["task_id"] = str(job_id)
    else:
        data["job_id"] = None
        data["task_id"] = None
    for key in ("config", "metrics", "ai_counts", "ai_pending"):
        parsed = _maybe_json(data.get(key))
        if parsed is None:
            parsed = {} if key in {"config", "metrics", "ai_counts"} else []
        data[key] = parsed
    data["total"] = int(data.get("total") or 0)
    data["processed"] = int(data.get("processed") or 0)
    rows_imported = data.get("rows_imported")
    data["rows_imported"] = int(rows_imported or data["processed"] or 0)
    data["imported"] = data["rows_imported"]
    if data["total"] and not data.get("pct"):
        try:
            pct = int(round((data["processed"] / max(data["total"], 1)) * 100))
            data["pct"] = max(0, min(100, pct))
        except Exception:
            pass
    data.setdefault("phase", data.get("phase") or "parse")
    data.setdefault("status", data.get("status") or "pending")
    data.setdefault("state", data.get("state") or data["status"])
    return data


def _get_import_status(task_id: str) -> Optional[Dict[str, Any]]:
    with _IMPORT_STATUS_LOCK:
        snapshot = dict(IMPORT_STATUS.get(task_id) or {})
    job_id = snapshot.get("job_id")
    if job_id is None:
        try:
            job_id = int(task_id)
        except Exception:
            job_id = None
    if job_id is not None:
        try:
            conn = ensure_db()
            row = database.get_import_job(conn, int(job_id))
        except Exception:
            row = None
        payload = _job_payload_from_row(row)
        if payload is not None:
            payload.update(snapshot)
            payload["job_id"] = payload.get("job_id") or job_id
            payload["task_id"] = payload.get("task_id") or (str(job_id) if job_id is not None else task_id)
            return payload
    if not snapshot:
        return None
    if job_id is not None:
        snapshot.setdefault("job_id", job_id)
        snapshot.setdefault("task_id", str(job_id))
    else:
        snapshot.setdefault("task_id", task_id)
    return snapshot


def _ensure_desire(product: Dict[str, Any], extras: Dict[str, Any]) -> str:
    """Return desire value from known sources.

    Precedence: product.desire -> extras.desire -> product.ai_desire ->
    product.ai_desire_label -> product.desire_magnitude.  Normalizes to
    string and logs when no value is found."""

    sources = [
        ("product.desire", rget(product, "desire")),
        ("extras.desire", rget(extras, "desire")),
        ("product.ai_desire", rget(product, "ai_desire")),
        ("product.ai_desire_label", rget(product, "ai_desire_label")),
        ("product.desire_magnitude", rget(product, "desire_magnitude")),
    ]
    desire_val = ""
    source_used = None
    for name, val in sources:
        if val not in (None, ""):
            desire_val = str(val)
            source_used = name
            break
    if desire_val == "":
        logger.info(
            "desire_missing=true sources_checked=%s product=%s",
            [s for s, _ in sources],
            rget(product, "id"),
        )
    else:
        logger.info(
            "product=%s desire=%s source=%s",
            rget(product, "id"),
            desire_val,
            source_used,
        )
    return desire_val


def parse_xlsx(binary: bytes):
    """Parse a minimal XLSX file into a list of dictionaries."""
    import zipfile
    import xml.etree.ElementTree as ET
    from io import BytesIO

    with zipfile.ZipFile(BytesIO(binary)) as z:
        shared = []
        if 'xl/sharedStrings.xml' in z.namelist():
            ss_root = ET.fromstring(z.read('xl/sharedStrings.xml'))
            for si in ss_root.findall('.//{*}si'):
                text = ''.join((t.text or '') for t in si.findall('.//{*}t'))
                shared.append(text)
        sheet_name = None
        for name in z.namelist():
            if name.startswith('xl/worksheets/sheet') and name.endswith('.xml'):
                sheet_name = name
                break
        if not sheet_name:
            return []
        root = ET.fromstring(z.read(sheet_name))
        rows = []
        for row in root.findall('.//{*}row'):
            values = []
            last_col_idx = 0
            for c in row.findall('{*}c'):
                cell_ref = c.attrib.get('r', '')
                letters = ''.join(ch for ch in cell_ref if ch.isalpha())
                col_idx = 0
                for ch in letters:
                    col_idx = col_idx * 26 + (ord(ch.upper()) - ord('A') + 1)
                while last_col_idx < col_idx - 1:
                    values.append('')
                    last_col_idx += 1
                val = ''
                cell_type = c.attrib.get('t')
                if cell_type == 's':
                    v = c.find('{*}v')
                    if v is not None:
                        try:
                            idx = int(v.text)
                            val = shared[idx] if idx < len(shared) else ''
                        except Exception:
                            val = ''
                elif cell_type == 'inlineStr':
                    tnode = c.find('{*}is/{*}t')
                    val = tnode.text if tnode is not None else ''
                else:
                    v = c.find('{*}v')
                    val = v.text if v is not None else ''
                values.append(val)
                last_col_idx = col_idx
            rows.append(values)
        while rows and all(not cell for cell in rows[0]):
            rows.pop(0)
        if not rows:
            return []
        headers = rows[0]
        records = []
        for r in rows[1:]:
            rec = {}
            for i, h in enumerate(headers):
                rec[h] = r[i] if i < len(r) else ''
            records.append(rec)
        return records


def _schedule_post_import_tasks(
    job_id: int,
    product_ids: List[int],
    rows_imported: int,
    task_key: str,
) -> None:
    auto_ai = config.is_auto_fill_ia_on_import_enabled()

    def status_cb(**payload: Any) -> None:
        payload.setdefault("job_id", job_id)
        payload.setdefault("task_id", task_key)
        _update_import_status(task_key, **payload)

    def worker() -> None:
        conn = database.get_connection(DB_PATH)
        try:
            final_counts: Dict[str, Any] = {
                "queued": len(product_ids),
                "sent": 0,
                "ok": 0,
                "ko": 0,
                "cached": 0,
                "retried": 0,
                "cost_spent_usd": 0.0,
            }
            pending_ids: List[int] = list(product_ids)
            try:
                if auto_ai and product_ids:
                    result = ai_columns.run_ai_fill_job(job_id, product_ids, status_cb=status_cb)
                    final_counts = result.get("counts", final_counts)
                    pending_ids = result.get("pending_ids", pending_ids)
                    error = result.get("error")
                    if error:
                        database.set_import_job_ai_error(conn, job_id, str(error))
                else:
                    final_counts = {
                        "queued": 0,
                        "sent": 0,
                        "ok": 0,
                        "ko": 0,
                        "cached": 0,
                        "retried": 0,
                        "cost_spent_usd": 0.0,
                    }
                    pending_ids = []
                    database.set_import_job_ai_counts(conn, job_id, final_counts, pending_ids)
                    database.update_import_job_ai_progress(conn, job_id, 0)
            except Exception as exc:
                logger.exception("AI enrichment failed job=%s", job_id)
                final_counts = {
                    "queued": len(product_ids),
                    "sent": 0,
                    "ok": 0,
                    "ko": len(product_ids),
                    "cached": 0,
                    "retried": 0,
                    "cost_spent_usd": 0.0,
                }
                pending_ids = list(product_ids)
                database.set_import_job_ai_error(conn, job_id, str(exc))
                database.set_import_job_ai_counts(conn, job_id, final_counts, pending_ids)
                database.update_import_job_ai_progress(conn, job_id, 0)
            total_ai = int(final_counts.get("queued", 0) or 0)
            done_ai = int(final_counts.get("ok", 0) + final_counts.get("cached", 0))
            _update_import_status(
                task_key,
                phase="winner",
                ai_counts=final_counts,
                ai_total=total_ai,
                ai_done=done_ai,
                ai_pending=pending_ids,
                message="Calculando Winner Score",
            )
            database.update_import_job_progress(conn, job_id, phase="winner")
            updated_scores = 0
            try:
                res_scores = winner_calc.generate_winner_scores(conn, product_ids=product_ids)
                updated_scores = int(res_scores.get("updated", 0))
            except Exception as exc:
                logger.exception("Winner score recalculation failed job=%s", job_id)
            database.complete_import_job(conn, job_id, rows_imported, updated_scores)
            _set_import_progress(
                task_key,
                pct=100,
                message="Completado",
                state="done",
                stage="done",
                imported=rows_imported,
                finished_at=time.time(),
                phase="done",
                ai_counts=final_counts,
                ai_total=total_ai,
                ai_done=done_ai,
                ai_pending=pending_ids,
            )
        finally:
            try:
                conn.close()
            except Exception:
                pass

    threading.Thread(target=worker, daemon=True).start()


def _process_import_job(job_id: int, tmp_path: Path, filename: str) -> None:
    """Background task to import XLSX data into the database."""
    conn = ensure_db()
    rows_imported = 0
    inserted_ids: List[int] = []
    task_key = str(job_id)
    try:
        _set_import_progress(
            task_key,
            pct=2,
            message="Inicializando importación",
            state="running",
            stage="prepare",
            started_at=time.time(),
            filename=filename,
        )
        data = tmp_path.read_bytes()
        records = parse_xlsx(data)
        total_records = len(records)
        if total_records:
            _set_import_progress(task_key, pct=5, message="Analizando archivo", total=total_records)

        used_cols: set[str] = set()

        def find_key(keys, patterns):
            for k in keys:
                if k in used_cols:
                    continue
                sanitized = ''.join(ch.lower() for ch in k if ch.isalnum())
                for p in patterns:
                    if p in sanitized:
                        used_cols.add(k)
                        return k
            return None

        if records:
            headers = list(records[0].keys())
            # identify columns with tolerant synonyms
            rating_col = find_key(headers, ["rating", "stars", "valoracion", "puntuacion"])
            units_col = find_key(headers, ["unitssold", "units", "ventas", "sold"])
            revenue_col = find_key(headers, ["revenue", "sales", "ingresos"])
            conv_col = find_key(headers, ["conversion", "cr", "tasaconversion"])
            launch_col = find_key(headers, ["launchdate", "fecha", "date", "firstseen"])
            range_col = None
            if "Date Range" in headers:
                range_col = "Date Range"
                used_cols.add(range_col)
            else:
                range_col = find_key(headers, ["daterange", "fecharango", "rangofechas"])
            price_col = find_key(headers, ["price", "precio", "cost", "unitprice"])
            img_col = find_key(headers, ["imageurl", "image", "imagelink", "mainimage", "mainimageurl", "img", "imagen", "picture", "primaryimage"])
            name_col = find_key(headers, ["name", "productname", "title", "product", "producto"])
            desc_col = find_key(headers, ["description", "descripcion", "desc"])
            cat_col = find_key(headers, ["category", "categoria", "niche", "segment"])
            curr_col = find_key(headers, ["currency", "moneda"])

            metric_names = [
                "magnitud_deseo",
                "nivel_consciencia_headroom",
                "evidencia_demanda",
                "tasa_conversion",
                "ventas_por_dia",
                "recencia_lanzamiento",
                "competition_level_invertido",
                "facilidad_anuncio",
                "escalabilidad",
                "durabilidad_recurrencia",
            ]

            def sanitize(name: str) -> str:
                return "".join(ch for ch in name if ch.isalnum())

            metric_cols = {m: find_key(headers, [sanitize(m)]) for m in metric_names}

            def parse_number(val: Any) -> Optional[float]:
                if val in (None, ''):
                    return None
                s = str(val).strip()
                if not s:
                    return None
                percent = '%' in s
                s = s.replace('%', '').replace(' ', '').replace(',', '.')
                s = re.sub(r'[^0-9.+-]', '', s)
                try:
                    num = float(s)
                    if percent:
                        num /= 100.0
                    return num
                except Exception:
                    return None

            def parse_text(val: Any) -> Optional[str]:
                if val is None:
                    return None
                s = str(val).strip()
                return s or None

            numeric_metrics = {
                "evidencia_demanda",
                "tasa_conversion",
                "ventas_por_dia",
                "recencia_lanzamiento",
            }

            cur = conn.cursor()
            cur.execute("BEGIN")
            cur.execute("SELECT COUNT(*) FROM products")
            count = cur.fetchone()[0]
            cur.execute("SELECT COALESCE(MAX(id), -1) FROM products")
            max_id = cur.fetchone()[0]
            is_empty = count == 0
            base_id = 0 if is_empty else (max_id + 1)
            rows_validas = []
            for row in records:
                name = (row.get(name_col) or '').strip() if name_col else None
                if not name:
                    continue
                description = (row.get(desc_col) or '').strip() if desc_col else None
                category = (row.get(cat_col) or '').strip() if cat_col else None
                price = None
                if price_col and row.get(price_col):
                    try:
                        price = float(str(row.get(price_col)).replace(',', '.'))
                    except Exception:
                        price = None
                currency = (row.get(curr_col) or '').strip() if curr_col else None
                image_url = (row.get(img_col) or '').strip() if img_col else None

                date_range = (row.get(range_col) or '').strip() if range_col else ''

                extras = {}
                metrics: dict[str, object] = {}

                rating_val = None
                if rating_col and row.get(rating_col) not in (None, ''):
                    try:
                        s = str(row.get(rating_col)).strip().replace(' ', '').replace(',', '.')
                        s = re.sub(r'[^0-9.]+', '', s)
                        if s.count('.') > 1:
                            parts = s.split('.')
                            s = ''.join(parts[:-1]) + '.' + parts[-1]
                        rating_val = float(s) if s else None
                    except Exception:
                        rating_val = None
                if rating_val is not None:
                    extras['rating'] = rating_val
                    extras['Product Rating'] = rating_val

                units_val = None
                if units_col and row.get(units_col) not in (None, ''):
                    try:
                        s = re.sub(r'[^0-9]+', '', str(row.get(units_col)))
                        units_val = int(s) if s else None
                    except Exception:
                        units_val = None
                if units_val is not None:
                    extras['units_sold'] = units_val
                    extras['Item Sold'] = units_val

                revenue_val = None
                if revenue_col and row.get(revenue_col) not in (None, ''):
                    try:
                        s = str(row.get(revenue_col)).strip().replace(' ', '').replace(',', '.')
                        s = re.sub(r'[^0-9.]+', '', s)
                        if s.count('.') > 1:
                            parts = s.split('.')
                            s = ''.join(parts[:-1]) + '.' + parts[-1]
                        revenue_val = float(s) if s else None
                    except Exception:
                        revenue_val = None
                if revenue_val is None and price is not None and units_val is not None:
                    revenue_val = price * units_val
                if revenue_val is not None:
                    extras['revenue'] = revenue_val
                    extras['Revenue($)'] = revenue_val

                def set_extra(col, key):
                    val = row.get(col) if col else None
                    if val is not None and str(val).strip():
                        extras[key] = str(val).strip()

                set_extra(conv_col, 'conversion_rate')
                set_extra(launch_col, 'launch_date')

                for m in metric_names:
                    col = metric_cols.get(m)
                    raw = row.get(col) if col else None
                    if raw not in (None, ''):
                        if m in numeric_metrics:
                            val = parse_number(raw)
                        else:
                            val = parse_text(raw)
                        if val is not None:
                            metrics[m] = val

                recognized = {
                    name_col,
                    desc_col,
                    cat_col,
                    price_col,
                    curr_col,
                    img_col,
                    rating_col,
                    units_col,
                    revenue_col,
                    conv_col,
                    launch_col,
                    range_col,
                }
                recognized.update(c for c in metric_cols.values() if c)
                for k, v in row.items():
                    if k not in recognized:
                        extras[k] = v

                rows_validas.append(
                    (name, description, category, price, currency, image_url, date_range, extras, metrics)
                )
            total_valid = len(rows_validas)
            if total_valid:
                _set_import_progress(
                    task_key,
                    pct=20,
                    message="Preparando inserción",
                    done=0,
                    total=total_valid,
                )
            for idx, (name, description, category, price, currency, image_url, date_range, extra_cols, metrics) in enumerate(rows_validas, start=1):
                row_id = base_id + idx
                database.insert_product(
                    conn,
                    name=name,
                    description=description,
                    category=category,
                    price=price,
                    currency=currency,
                    image_url=image_url,
                    date_range=date_range,
                    source=filename,
                    extra=extra_cols,
                    commit=False,
                    product_id=row_id,
                )
                if metrics:
                    database.update_product(conn, row_id, **metrics)
                rows_imported += 1
                inserted_ids.append(row_id)
                if total_valid:
                    if idx == total_valid or idx % 50 == 0:
                        frac = idx / max(total_valid, 1)
                        pct = 20 + min(60, 60 * frac)
                        _set_import_progress(
                            task_key,
                            pct=pct,
                            message=f"Insertando registros {idx}/{total_valid}",
                            done=idx,
                            total=total_valid,
                        )
            conn.commit()
            _set_import_progress(
                task_key,
                pct=82,
                message="Guardando cambios",
                done=rows_imported,
                total=total_valid or rows_imported,
            )

        next_phase = "enrich" if inserted_ids and config.is_auto_fill_ia_on_import_enabled() else "winner"
        database.update_import_job_progress(
            conn,
            job_id,
            phase=next_phase,
            status="running",
            processed=rows_imported,
            total=rows_imported,
            rows_imported=rows_imported,
        )
        _set_import_progress(
            task_key,
            pct=90,
            message="Importación completada",
            state="done",
            stage="done",
            imported=rows_imported,
            done=rows_imported,
            total=rows_imported,
            phase=next_phase,
        )
        _schedule_post_import_tasks(job_id, inserted_ids, rows_imported, task_key)
    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        database.fail_import_job(conn, job_id, str(exc))
        _set_import_progress(
            task_key,
            pct=100,
            message=f"Error: {exc}",
            state="error",
            stage="error",
            error=str(exc),
            finished_at=time.time(),
        )
    finally:
        try:
            tmp_path.unlink()
        except Exception:
            pass


def resume_incomplete_imports():
    """Mark stale pending imports as failed and remove orphan temp files."""
    conn = ensure_db()
    database.mark_stale_pending_imports(conn, 5)
    tmp_dir = APP_DIR / 'uploads'
    if tmp_dir.exists():
        cur = conn.cursor()
        cur.execute("SELECT temp_path FROM import_jobs")
        valid = {Path(row[0]) for row in cur.fetchall() if row[0]}
        for f in tmp_dir.glob('import_*'):
            if f not in valid:
                try:
                    f.unlink()
                except Exception:
                    pass


def _start_enrichment_worker(job_id: int) -> bool:
    def runner() -> None:
        try:
            product_enrichment.run_job_sync(job_id)
        except Exception:
            logger.exception("Enrichment worker crashed job_id=%s", job_id)
        finally:
            with _ENRICH_LOCK:
                _ENRICH_WORKERS.pop(job_id, None)

    with _ENRICH_LOCK:
        if job_id in _ENRICH_WORKERS:
            return False
        thread = threading.Thread(target=runner, name=f"enrich-{job_id}", daemon=True)
        _ENRICH_WORKERS[job_id] = thread
        thread.start()
        return True


class _SilentWriter:
    """Wrapper around a socket writer that ignores connection errors."""

    def __init__(self, raw):
        self._raw = raw

    def write(self, data):
        try:
            self._raw.write(data)
        except (BrokenPipeError, ConnectionResetError, ConnectionAbortedError):
            pass

    def flush(self):
        try:
            self._raw.flush()
        except (BrokenPipeError, ConnectionResetError, ConnectionAbortedError):
            pass

    def __getattr__(self, name):
        return getattr(self._raw, name)


class RequestHandler(BaseHTTPRequestHandler):
    server_version = "ProductResearchCopilot/1.0"

    def setup(self):
        super().setup()
        self.wfile = _SilentWriter(self.wfile)

    def _set_json(self, status=200):
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PATCH, OPTIONS")
        self.send_header(
            "Cache-Control", "no-store, no-cache, must-revalidate, max-age=0"
        )
        self.end_headers()

    def _set_html(self, status=200):
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PATCH, OPTIONS")
        self.end_headers()

    def safe_write(self, func):
        try:
            func()
            return True
        except (BrokenPipeError, ConnectionResetError, ConnectionAbortedError):
            return False

    def send_json(self, obj: Any, status: int = 200):
        self._set_json(status)
        self.wfile.write(json.dumps(obj).encode('utf-8'))

    def _safe_write(self, data: bytes) -> bool:
        return self.safe_write(lambda: self.wfile.write(data))

    def _serve_static(self, rel_path: str):
        file_path = STATIC_DIR / rel_path
        if not file_path.exists():
            self.send_error(404)
            return
        if file_path.suffix == ".js":
            ctype = "application/javascript"
        elif file_path.suffix == ".css":
            ctype = "text/css"
        elif file_path.suffix == ".html":
            ctype = "text/html"
        else:
            ctype = "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", f"{ctype}; charset=utf-8")
        self.end_headers()
        with open(file_path, "rb") as f:
            self.wfile.write(f.read())

    def _parse_multipart_file(self):
        ctype = self.headers.get('Content-Type', '')
        if not ctype.startswith('multipart/form-data'):
            return None, None
        boundary_key = 'boundary='
        if boundary_key not in ctype:
            return None, None
        boundary = ctype.split(boundary_key, 1)[1].strip().strip('"')
        try:
            length = int(self.headers.get('Content-Length', '0'))
        except ValueError:
            length = 0
        if length <= 0:
            return None, None
        body = self.rfile.read(length)
        parser = BytesParser(policy=default)
        header_bytes = f'Content-Type: multipart/form-data; boundary={boundary}\r\n\r\n'.encode('utf-8')
        msg = parser.parsebytes(header_bytes + body)
        for part in msg.iter_parts():
            disp = part.get_content_disposition()
            if disp == 'form-data' and part.get_param('name', header='content-disposition') == 'file':
                filename = part.get_filename()
                data = part.get_payload(decode=True)
                return filename, data
        return None, None

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PATCH, OPTIONS")
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/" or path == "/index.html":
            self._serve_static("index.html")
            return
        if path.startswith("/static/"):
            rel = path[len("/static/") :]
            self._serve_static(rel)
            return
        if path == "/api/log-path":
            self._set_json()
            self.wfile.write(json.dumps({"path": str(LOG_PATH)}).encode("utf-8"))
            return
        if path == "/api/auth/has-key":
            has_key = bool(config.get_api_key())
            self._set_json()
            self.wfile.write(json.dumps({"ok": True, "has_key": has_key}).encode("utf-8"))
            return
        if path == "/api/trends/summary":
            params = parse_qs(parsed.query)
            qs_from = params.get("from", [""])[0]
            qs_to = params.get("to", [""])[0]
            filters_s = params.get("filters", [None])[0]

            today = date.today()
            d_from = _parse_date(qs_from)
            d_to = _parse_date(qs_to)

            if d_from is None and d_to is None:
                d_to = today
                d_from = today - timedelta(days=29)
            elif d_from is None:
                d_from = d_to - timedelta(days=29)
            elif d_to is None:
                d_to = d_from + timedelta(days=29)

            if d_from > d_to:
                d_from, d_to = d_to, d_from

            start_dt = datetime.combine(d_from, datetime.min.time())
            end_dt = datetime.combine(d_to + timedelta(days=1), datetime.min.time())

            filters = None
            if filters_s:
                try:
                    filters = json.loads(filters_s)
                except Exception:
                    filters = None
            try:
                resp = trends_service.get_trends_summary(start_dt, end_dt, filters)
            except Exception:
                resp = {
                    "categories": [],
                    "timeseries": [],
                    "granularity": "day",
                    "totals": {
                        "unique_products": 0,
                        "units": 0,
                        "revenue": 0,
                        "avg_price": 0,
                        "avg_rating": 0,
                        "rev_per_unit": 0,
                        "delta_revenue_pct": 0,
                        "delta_units_pct": 0,
                    },
                }
            self._set_json()
            self.wfile.write(json.dumps(resp).encode("utf-8"))
            return
        if path == "/api/config/winner-weights":
            cfg = config.load_config()
            changed = False
            order = cfg.get("winner_order")
            if not isinstance(order, list) or len(order) != 8:
                order = list(config.DEFAULT_WINNER_ORDER)
                cfg["winner_order"] = order[:]
                cfg["weights_order"] = order[:]
                changed = True
            weights_order = cfg.get("weights_order")
            if not isinstance(weights_order, list) or len(weights_order) != 8:
                cfg["weights_order"] = cfg["winner_order"][:]
                weights_order = cfg["weights_order"]
                changed = True

            weights_map = _sanitize_weights_map(cfg.get("winner_weights"))
            weights_enabled = _sanitize_enabled_map(
                cfg.get("weights_enabled"),
                list(weights_map.keys()),
            )
            if weights_enabled != cfg.get("weights_enabled"):
                cfg["weights_enabled"] = weights_enabled
                changed = True

            order = _normalize_order_list(cfg.get("weights_order"), weights_map)
            if order != cfg.get("winner_order"):
                cfg["winner_order"] = order[:]
                cfg["weights_order"] = order[:]
                changed = True

            if changed:
                config.save_config(cfg)

            logger.info("CONFIG served weights_order=%s", cfg.get("weights_order"))
            
            raw_eff = {
                k: (weights_map.get(k, 0) if weights_enabled.get(k, True) else 0)
                for k in weights_map
            }
            eff_int = winner_calc.compute_effective_int(raw_eff, order)
            logger.info("weights_effective_int=%s order=%s", eff_int, order)
            resp = {
                **weights_map,
                "weights": weights_map,
                "order": order,
                "effective": {"int": eff_int},
                "weights_enabled": weights_enabled,
                "weights_order": order,
            }
            self._set_json()
            self.wfile.write(json.dumps(resp).encode("utf-8"))
            return
        if path == "/_import_history":
            params = parse_qs(parsed.query)
            try:
                limit = int(params.get("limit", ["20"])[0])
            except Exception:
                limit = 20
            conn = ensure_db()
            rows = [row_to_dict(r) for r in database.get_import_history(conn, limit)]
            self.safe_write(lambda: self.send_json(rows))
            return
        if path in {"/_import_status", "/import/status"}:
            params = parse_qs(parsed.query)
            target = ""
            if path == "/import/status":
                target = params.get("job_id", [""])[0] or params.get("task_id", [""])[0]
            else:
                target = params.get("task_id", [""])[0] or params.get("job_id", [""])[0]
            if not target:
                self.safe_write(lambda: self.send_json({"state": "unknown"}))
                return
            status = _get_import_status(str(target))
            if status is None and str(target).isdigit():
                conn = ensure_db()
                row = database.get_import_job(conn, int(target))
                status = _job_payload_from_row(row)
            if status:
                status.setdefault("task_id", str(status.get("task_id") or target))
                if str(target).isdigit():
                    status.setdefault("job_id", int(target))
                self.safe_write(lambda: self.send_json(status))
            else:
                self.safe_write(lambda: self.send_json({"state": "unknown"}))
            return
        if path == "/enrich/status":
            params = parse_qs(parsed.query)
            job_raw = params.get("job_id", [""])[0]
            try:
                job_id = int(job_raw)
            except (TypeError, ValueError):
                self.safe_write(lambda: self.send_json({"error": "invalid_job_id"}, status=400))
                return
            conn = ensure_db()
            payload = database.get_enrichment_status(conn, job_id)
            if payload is None:
                self.safe_write(lambda: self.send_json({"error": "job_not_found"}, status=404))
                return
            self.safe_write(lambda: self.send_json(payload))
            return
        if path == "/metrics":
            params = parse_qs(parsed.query)
            try:
                limit = int(params.get("limit", ["20"])[0])
            except Exception:
                limit = 20
            conn = ensure_db()
            batches = [
                {
                    "job_id": row["job_id"],
                    "batch": row["batch_no"],
                    "rows": row["rows"],
                    "duration_ms": row["duration_ms"],
                    "throughput": row["throughput"],
                    "created_at": row["created_at"],
                }
                for row in database.get_recent_import_metrics(conn, limit)
            ]
            ai_batches = [
                {
                    "job_id": row["job_id"],
                    "batch": row["batch_no"],
                    "rows": row["rows"],
                    "duration_ms": row["duration_ms"],
                    "throughput_rps": row["throughput"],
                    "cached_hits": row["cached_hits"],
                    "created_at": row["created_at"],
                }
                for row in database.get_recent_ai_metrics(conn, limit)
            ]
            jobs_payload = []
            for row in database.get_import_history(conn, limit):
                payload = _job_payload_from_row(row)
                if payload:
                    jobs_payload.append(payload)
            config_payload = {
                "pragmas": get_last_performance_config(),
                "default_batch_size": DEFAULT_BATCH_SIZE,
            }
            self.safe_write(
                lambda: self.send_json(
                    {
                        "jobs": jobs_payload,
                        "batches": batches,
                        "ai_batches": ai_batches,
                        "config": config_payload,
                    }
                )
            )
            return
        if path in ("/products", "/api/products"):
            # Return a list of products including extra metadata for UI display
            conn = ensure_db()
            rows = []
            for p_row in database.list_products(conn):
                p = row_to_dict(p_row)
                extra = rget(p, "extra", {})
                try:
                    extra_dict = json.loads(extra) if isinstance(extra, str) else (extra or {})
                except Exception:
                    extra_dict = {}
                if 'rating' in extra_dict and 'Product Rating' not in extra_dict:
                    extra_dict['Product Rating'] = extra_dict['rating']
                if 'units_sold' in extra_dict and 'Item Sold' not in extra_dict:
                    extra_dict['Item Sold'] = extra_dict['units_sold']
                if 'revenue' in extra_dict and 'Revenue($)' not in extra_dict:
                    extra_dict['Revenue($)'] = extra_dict['revenue']
                score_value = rget(p, "winner_score")
                dr = rget(p, "date_range")
                if dr is None:
                    dr = extra_dict.get("date_range")
                price_val = rget(p, "price")
                desire_db = rget(p, "desire")
                if desire_db in (None, ""):
                    desire_db = _ensure_desire(p, extra_dict)
                desire_val = (desire_db or "").strip() or None
                row = {
                    "id": rget(p, "id"),
                    "name": sanitize_product_name(rget(p, "name")),
                    "category": rget(p, "category"),
                    "price": price_val,
                    "image_url": rget(p, "image_url"),
                    "desire": desire_val,
                    "desire_magnitude": rget(p, "desire_magnitude"),
                    "awareness_level": rget(p, "awareness_level"),
                    "competition_level": rget(p, "competition_level"),
                    "rating": extra_dict.get("rating"),
                    "units_sold": extra_dict.get("units_sold"),
                    "revenue": extra_dict.get("revenue"),
                    "conversion_rate": extra_dict.get("conversion_rate"),
                    "launch_date": extra_dict.get("launch_date"),
                    "date_range": dr or "",
                    "extras": extra_dict,
                }
                if price_val is not None:
                    try:
                        row["price_display"] = round(float(price_val), 2)
                    except Exception:
                        row["price_display"] = price_val
                row["winner_score"] = score_value
                rows.append(row)
            self._set_json()
            self.wfile.write(json.dumps(rows).encode("utf-8"))
            return
        if path == "/config":
            # return stored configuration (without exposing the API key)
            cfg = config.load_config()
            changed = False
            order = cfg.get("winner_order")
            if not isinstance(order, list) or len(order) != 8:
                order = list(config.DEFAULT_WINNER_ORDER)
                cfg["winner_order"] = order[:]
                cfg["weights_order"] = order[:]
                changed = True
            if not isinstance(cfg.get("weights_order"), list) or len(cfg["weights_order"]) != 8:
                cfg["weights_order"] = cfg["winner_order"][:]
                changed = True
            if changed:
                config.save_config(cfg)

            weights_map = _sanitize_weights_map(cfg.get("winner_weights"))
            weights_enabled = _sanitize_enabled_map(
                cfg.get("weights_enabled"),
                list(weights_map.keys()),
            )
            key = cfg.get("api_key") or ""
            data = {
                "model": cfg.get("model", "gpt-4o"),
                "weights": weights_map,
                "winner_weights": weights_map,
                "winner_order": list(cfg.get("winner_order", list(DEFAULT_ORDER_LIST))),
                "weights_order": list(cfg.get("weights_order", list(DEFAULT_ORDER_LIST))),
                "weights_enabled": weights_enabled,
                "order": list(cfg.get("winner_order", list(DEFAULT_ORDER_LIST))),
                "has_api_key": bool(key),
                "oldness_preference": cfg.get("oldness_preference", "newer"),
                "weightsUpdatedAt": cfg.get("weightsUpdatedAt", 0),
                "weightsVersion": cfg.get("weightsVersion", 0),
            }
            if key:
                data["api_key_last4"] = key[-4:]
                data["api_key_length"] = len(key)
                data["api_key_hash"] = hashlib.sha256(key.encode("utf-8")).hexdigest()
            logger.info("CONFIG served weights_order=%s", data["weights_order"])
            self._set_json()
            self.wfile.write(json.dumps(data).encode("utf-8"))
            return
        if path == "/settings/winner-score":
            cfg = config.load_config()
            weights = cfg.get("weights", {})
            self._set_json()
            self.wfile.write(json.dumps(weights).encode("utf-8"))
            return
        if path == "/api/winner-score/explain" and DEBUG:
            self.handle_winner_score_explain(parsed)
            return
        if path.startswith("/score/"):
            try:
                pid = int(path.split("/")[-1])
            except ValueError:
                self.send_error(400, "Invalid product ID")
                return
            conn = ensure_db()
            scores = database.get_scores_for_product(conn, pid)
            if not scores:
                self._set_json(404)
                self.wfile.write(json.dumps({"error": "No score"}).encode("utf-8"))
                return
            score = row_to_dict(scores[0])
            self._set_json()
            self.wfile.write(json.dumps(score).encode("utf-8"))
            return
        if path == "/lists":
            # return all saved groups/lists with product counts
            conn = ensure_db()
            lsts = database.get_lists(conn)
            data = []
            for l in lsts:
                data.append({"id": l["id"], "name": l["name"], "count": l["count"]})
            self._set_json()
            self.wfile.write(json.dumps(data).encode("utf-8"))
            return
        if path.startswith("/list/"):
            # return products belonging to a list
            parts = path.strip("/").split("/")
            if len(parts) == 2 and parts[0] == "list":
                try:
                    lid = int(parts[1])
                except Exception:
                    self.send_error(400, "Invalid list ID")
                    return
                conn = ensure_db()
                prods = database.get_products_in_list(conn, lid)
                rows = []
                for p in prods:
                    scores = database.get_scores_for_product(conn, p["id"])
                    score = scores[0] if scores else None
                    extra = p["extra"] if "extra" in p.keys() else {}
                    try:
                        extra_dict = json.loads(extra) if isinstance(extra, str) else (extra or {})
                    except Exception:
                        extra_dict = {}
                    p_dict = row_to_dict(p)
                    desire_val = _ensure_desire(p_dict, extra_dict)
                    score_dict = row_to_dict(score)
                    score_value = rget(score_dict, "winner_score")
                    breakdown_data = {}
                    if score_dict:
                        try:
                            raw_breakdown = rget(score_dict, "winner_score_breakdown")
                            breakdown_data = json.loads(raw_breakdown or "{}")
                        except Exception:
                            breakdown_data = {}
                    row = {
                        "id": p["id"],
                        "name": sanitize_product_name(p["name"]),
                        "category": p["category"],
                        "price": p["price"],
                        "image_url": p["image_url"],
                        "desire": desire_val,
                        "desire_magnitude": rget(p_dict, "desire_magnitude"),
                        "extras": extra_dict,
                    }
                    row["winner_score"] = score_value
                    if score_dict:
                        row["winner_score_breakdown"] = breakdown_data
                    rows.append(row)
                self._set_json()
                self.wfile.write(json.dumps(rows).encode("utf-8"))
                return
        # trends endpoint: compute analytics for categories, keywords and scatter plots
        if path == "/trends":
            qs = parse_qs(parsed.query)
            start_str = qs.get("start", [None])[0]
            end_str = qs.get("end", [None])[0]
            def parse_date_str(val: Optional[str]):
                if not val:
                    return None
                for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
                    try:
                        return datetime.strptime(val, fmt)
                    except Exception:
                        continue
                return None

            start_dt = parse_date_str(start_str)
            end_dt = parse_date_str(end_str)

            conn = ensure_db()
            prods = database.list_products(conn)
            from collections import Counter, defaultdict

            cat_rev_growth: Dict[str, float] = defaultdict(float)
            cat_unit_growth: Dict[str, float] = defaultdict(float)
            cat_rev: Dict[str, float] = defaultdict(float)
            cat_units: Dict[str, float] = defaultdict(float)
            cat_product_count: Dict[str, int] = defaultdict(int)
            cat_price_total: Dict[str, float] = defaultdict(float)
            cat_price_count: Dict[str, int] = defaultdict(int)
            cat_rating_total: Dict[str, float] = defaultdict(float)
            cat_rating_count: Dict[str, int] = defaultdict(int)
            word_counter = Counter()
            brand_counter = Counter()
            scatter_rating_revenue = []
            scatter_price_revenue = []
            total_revenue = 0.0
            total_units = 0.0
            price_sum = 0.0
            price_count = 0
            top_product_name = None
            top_product_rev = 0.0

            stopwords = set([
                "the", "and", "for", "with", "a", "an", "de", "la", "el", "para", "y", "con", "un", "una", "los", "las", "en", "por", "to", "of",
            ])

            def parse_float(val):
                try:
                    return float(str(val).replace("%", "").replace("$", "").replace(",", "").strip())
                except Exception:
                    return None
            import re

            for p in prods:
                try:
                    extras = json.loads(p["extra"]) if p["extra"] else {}
                except Exception:
                    extras = {}

                launch_val = extras.get("Launch Date")
                launch_dt = parse_date_str(str(launch_val)) if launch_val else None
                if start_dt and (launch_dt is None or launch_dt < start_dt):
                    continue
                if end_dt and (launch_dt is None or launch_dt > end_dt):
                    continue
                cat = (p["category"] or "").strip().lower()
                if cat:
                    cat_product_count[cat] += 1
                rev_growth = None
                unit_growth = None
                for k, v in extras.items():
                    lk = k.lower()
                    if rev_growth is None and "revenue" in lk and "growth" in lk:
                        rev_growth = parse_float(v)
                    if unit_growth is None and ("item" in lk or "unit" in lk) and "growth" in lk:
                        unit_growth = parse_float(v)
                if rev_growth is not None and cat:
                    cat_rev_growth[cat] += rev_growth
                if unit_growth is not None and cat:
                    cat_unit_growth[cat] += unit_growth
                revenue = None
                for key in ["Revenue($)", "Revenue"]:
                    if key in extras:
                        revenue = parse_float(extras[key])
                        if revenue is not None:
                            break
                item_sold = parse_float(extras.get("Item Sold"))
                if revenue is not None:
                    total_revenue += revenue
                    if cat:
                        cat_rev[cat] += revenue
                    if revenue > top_product_rev:
                        top_product_rev = revenue
                        top_product_name = name_clean
                if item_sold is not None:
                    total_units += item_sold
                    if cat:
                        cat_units[cat] += item_sold
                name = name_clean.lower()
                words = re.split(r"[^a-záéíóúüñ0-9]+", name)
                for w in words:
                    if not w or w in stopwords or len(w) < 3:
                        continue
                    word_counter[w] += 1
                tokens = re.split(r"[^A-Za-z0-9]+", name_clean)
                if tokens:
                    brand = tokens[0].lower()
                    if brand and brand not in stopwords and len(brand) >= 3:
                        brand_counter[brand] += 1
                rating = parse_float(extras.get("Product Rating"))
                if rating is not None:
                    if revenue is not None and item_sold is not None:
                        scatter_rating_revenue.append({
                            "x": rating,
                            "y": revenue,
                            "r": item_sold,
                            "label": name_clean,
                            "units": item_sold,
                            "rating": rating,
                            "revenue": revenue,
                        })
                    if cat:
                        cat_rating_total[cat] += rating
                        cat_rating_count[cat] += 1
                avg_price = None
                for key in ["Avg. Unit Price($)", "Avg Unit Price($)", "Avg. Unit Price"]:
                    if key in extras:
                        avg_price = parse_float(extras[key])
                        if avg_price is not None:
                            break
                if avg_price is not None:
                    price_sum += avg_price
                    price_count += 1
                    if cat:
                        cat_price_total[cat] += avg_price
                        cat_price_count[cat] += 1
                    if revenue is not None:
                        scatter_price_revenue.append({
                            "x": avg_price,
                            "y": revenue,
                            "label": name_clean,
                            "units": item_sold,
                            "rating": rating,
                            "revenue": revenue,
                        })

            cat_rev_per_unit = []
            for cat, rev in cat_rev.items():
                units = cat_units.get(cat, 0)
                if units:
                    cat_rev_per_unit.append((cat, rev / units))

            top_rev_growth = sorted(cat_rev_growth.items(), key=lambda x: x[1], reverse=True)[:10]
            top_unit_growth = sorted(cat_unit_growth.items(), key=lambda x: x[1], reverse=True)[:10]
            cat_rev_per_unit.sort(key=lambda x: x[1], reverse=True)
            top_words = word_counter.most_common(10)
            top_brands = [(b.title(), c) for b, c in brand_counter.most_common(10)]
            avg_price = price_sum / price_count if price_count else 0.0
            top_cat = None
            if cat_rev:
                top_cat = max(cat_rev.items(), key=lambda x: x[1])[0]
            category_compare = []
            category_summary = []
            for cat, count in cat_product_count.items():
                total_r = cat_rev.get(cat, 0.0)
                total_u = cat_units.get(cat, 0.0)
                avg_rev = total_r / count if count else 0.0
                avg_units = total_u / count if count else 0.0
                avg_p = cat_price_total[cat] / cat_price_count[cat] if cat_price_count[cat] else 0.0
                avg_r = cat_rating_total[cat] / cat_rating_count[cat] if cat_rating_count[cat] else 0.0
                category_compare.append({
                    "category": cat.title(),
                    "products": count,
                    "avg_revenue": avg_rev,
                    "avg_units": avg_units,
                    "avg_price": avg_p,
                    "total_revenue": total_r,
                    "total_units": total_u,
                    "avg_rating": avg_r,
                })
                
                category_summary.append({
                    "category": cat.title(),
                    "products": count,
                    "total_units": total_u,
                    "total_revenue": total_r,
                    "avg_price": avg_p,
                    "avg_rating": avg_r,
                })

            rows = []
            for p in prods:
                scores = database.get_scores_for_product(conn, p["id"])
                score_val = None
                if scores:
                    try:
                        score_val = scores[0]["winner_score"]
                    except Exception:
                        score_val = None
                if score_val is not None:
                    clean_name = sanitize_product_name(p["name"]) or ""
                    rows.append((p["id"], clean_name, score_val))
            rows.sort(key=lambda x: x[2], reverse=True)
            key_name = "winner_score"
            top_products = [{"id": r[0], "name": r[1], "winner_score": r[2]} for r in rows[:10]]

            self._set_json()
            self.wfile.write(json.dumps({
                "kpis": {
                    "total_revenue": total_revenue,
                    "total_units": total_units,
                    "avg_price": avg_price,
                    "top_category": top_cat.title() if top_cat else None,
                    "top_product": top_product_name,
                },
                "category_compare": category_compare,
                "category_summary": category_summary,
                "cat_revenue_growth": top_rev_growth,
                "cat_units_growth": top_unit_growth,
                "cat_rev_per_unit": cat_rev_per_unit[:10],
                "keywords": top_words,
                "brands": top_brands,
                "top_products": top_products,
                "scatter_rating_revenue": scatter_rating_revenue,
                "scatter_price_revenue": scatter_price_revenue,
            }).encode("utf-8"))
            return
# export selected or all products
        if path == "/export":
            qs = parse_qs(parsed.query)
            fmt = qs.get('format', ['csv'])[0]
            id_str = qs.get('ids', [None])[0]
            conn = ensure_db()
            items: list[sqlite3.Row] = []
            if id_str:
                try:
                    ids = [int(x) for x in id_str.split(',') if x]
                except Exception:
                    ids = []
                for pid in ids:
                    p = database.get_product(conn, pid)
                    if p:
                        items.append(p)
            else:
                items = database.list_products(conn)
            rows = []
            for p in items:
                scores = database.get_scores_for_product(conn, p['id'])
                score_val = None
                if scores:
                    sc = scores[0]
                    if 'winner_score' in sc.keys():
                        score_val = sc['winner_score']
                rows.append(
                    [
                        p['id'],
                        sanitize_product_name(p['name']) or "",
                        score_val,
                        p['desire'],
                        p['desire_magnitude'],
                        p['awareness_level'],
                        p['competition_level'],
                        p['date_range'],
                    ]
                )
            headers = ["id", "name", "Winner Score", "Desire", "Desire Magnitude", "Awareness Level", "Competition Level", "Date Range"]
            if fmt == 'xlsx':
                try:
                    from openpyxl import Workbook
                except Exception:
                    self._set_json(500)
                    self.wfile.write(json.dumps({"error": "openpyxl not installed"}).encode('utf-8'))
                    return
                wb = Workbook()
                ws = wb.active
                ws.append(headers)
                for r in rows:
                    ws.append(r)
                from io import BytesIO
                bio = BytesIO()
                wb.save(bio)
                data = bio.getvalue()
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", "attachment; filename=export.xlsx")
                self.end_headers()
                self.wfile.write(data)
                return
            else:
                import csv
                from io import StringIO
                output = StringIO()
                writer = csv.writer(output)
                writer.writerow(headers)
                writer.writerows(rows)
                csv_data = output.getvalue().encode('utf-8')
                self.send_response(200)
                self.send_header("Content-Type", "text/csv; charset=utf-8")
                self.send_header("Content-Disposition", "attachment; filename=export.csv")
                self.end_headers()
                self.wfile.write(csv_data)
                return
        # unknown
        self.send_error(404)
        # unknown
        self.send_error(404)

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path.startswith("/api/gpt/"):
            task = path[len("/api/gpt/") :]
            if not task:
                self._set_json(404)
                self.wfile.write(json.dumps({"error": "unknown_task"}).encode('utf-8'))
                return
            self.handle_prompt_task(task)
            return
        if path == "/api/analyze/titles":
            self.handle_analyze_titles()
            return
        if path == "/api/export/kalodata-minimal":
            self.handle_export_kalodata_minimal()
            return
        if path == "/api/auth/set-key":
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length).decode('utf-8')
            try:
                data = json.loads(body)
                key = str(data.get("api_key", "")).strip()
            except Exception:
                self._set_json(400)
                self.wfile.write(json.dumps({"ok": False, "has_key": False, "error": "invalid_json"}).encode('utf-8'))
                return
            if not key:
                self._set_json(400)
                self.wfile.write(json.dumps({"ok": False, "has_key": False, "error": "empty_api_key"}).encode('utf-8'))
                return
            try:
                resp = requests.get(
                    "https://api.openai.com/v1/models",
                    headers={"Authorization": f"Bearer {key}"},
                    timeout=15,
                )
                if resp.status_code != 200:
                    raise ValueError(resp.text)
            except Exception as exc:
                self._set_json(400)
                self.wfile.write(json.dumps({"ok": False, "has_key": False, "error": str(exc)}).encode('utf-8'))
                return
            cfg = config.load_config()
            cfg["api_key"] = key
            config.save_config(cfg)
            self._set_json()
            self.wfile.write(json.dumps({"ok": True, "has_key": True}).encode('utf-8'))
            return
        if path == "/upload":
            self.handle_upload()
            return
        if path == "/evaluate_all":
            self.handle_evaluate_all()
            return
        if path == "/setconfig":
            self.handle_setconfig()
            return
        if path == "/custom_gpt":
            self.handle_custom_gpt()
            return
        if path == "/api/ba/insights":
            self.handle_ba_insights()
            return
        if path == "/api/ia/batch-columns":
            self.handle_ia_batch_columns()
            return
        if path == "/auto_weights":
            self.handle_auto_weights()
            return
        if path == "/api/config/winner-weights/ai":
            self.handle_scoring_v2_auto_weights_gpt()
            return
        if path == "/api/config/winner-weights/reset":
            cfg = _apply_weights_reset()
            payload = _build_weights_payload(cfg)
            logger.info("RESET applied weights_order=%s", cfg.get("weights_order"))
            self._set_json()
            self.wfile.write(json.dumps(payload).encode('utf-8'))
            return
        if path == "/scoring/v2/auto-weights-gpt":
            self.handle_scoring_v2_auto_weights_gpt()
            return
        if path == "/scoring/v2/auto-weights-stat":
            self.handle_scoring_v2_auto_weights_stat()
            return
        if path == "/scoring/v2/gpt-evaluate":
            self.handle_scoring_v2_gpt_evaluate()
            return
        if path == "/scoring/v2/gpt-summary":
            self.handle_scoring_v2_gpt_summary()
            return
        if path == "/scoring/v2/generate" or path == "/api/winner-score/generate":
            self.handle_scoring_v2_generate()
            return
        if path == "/delete":
            self.handle_delete()
            return
        if path == "/remove_from_list":
            self.handle_remove_from_list()
            return
        if path == "/create_list":
            self.handle_create_list()
            return
        if path == "/delete_list":
            self.handle_delete_list()
            return
        if path == "/add_to_list":
            self.handle_add_to_list()
            return
        if path == "/shutdown":
            self.handle_shutdown()
            return
        if path == "/enrich/start":
            params = parse_qs(parsed.query)
            job_raw = params.get("job_id", [""])[0]
            try:
                job_id = int(job_raw)
            except (TypeError, ValueError):
                self.safe_write(lambda: self.send_json({"error": "invalid_job_id"}, status=400))
                return
            conn = ensure_db()
            job = database.get_import_job(conn, job_id)
            if job is None:
                self.safe_write(lambda: self.send_json({"error": "job_not_found"}, status=404))
                return
            config_data = product_enrichment.parse_job_config(job["config"])
            full_config, _ = product_enrichment.ensure_enrich_config(config_data)
            database.update_import_job_progress(
                conn,
                job_id,
                phase="enrich",
                status="enriching",
                config=full_config,
            )
            started = _start_enrichment_worker(job_id)
            payload = database.get_enrichment_status(conn, job_id) or {
                "job_id": job_id,
                "phase": "enrich",
            }
            payload["started"] = started
            publish_progress({"event": "enrich", **payload})
            logger.info("enrich start job=%s started=%s", job_id, started)
            self.safe_write(lambda: self.send_json(payload))
            return
        if path == "/products":
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length).decode('utf-8')
            try:
                data = json.loads(body)
                if not isinstance(data, dict):
                    raise ValueError
            except Exception:
                self._set_json(400)
                self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
                return
            if "price" in data and data.get("source") != "import":
                logger.info(
                    "price field is read-only; ignoring on create (source=%s)",
                    data.get("source"),
                )
                data.pop("price", None)
            conn = ensure_db()
            pid = database.insert_product(
                conn,
                name=data.get("name", ""),
                description=data.get("description"),
                category=data.get("category"),
                price=data.get("price"),
                currency=data.get("currency"),
                image_url=data.get("image_url"),
                source=data.get("source"),
                desire=data.get("desire"),
                desire_magnitude=data.get("desire_magnitude"),
                awareness_level=data.get("awareness_level"),
                competition_level=data.get("competition_level"),
                extra=data.get("extras"),
            )
            product = row_to_dict(database.get_product(conn, pid))
            try:
                extra_dict = json.loads(rget(product, "extra") or "{}")
            except Exception:
                extra_dict = {}
            product["desire"] = _ensure_desire(product, extra_dict)
            self._set_json()
            self.wfile.write(json.dumps(product).encode('utf-8'))
            return
        self.send_error(404)

    def do_PUT(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path.startswith("/products/"):
            try:
                pid = int(path.split("/")[-1])
            except Exception:
                self._set_json(400)
                self.wfile.write(json.dumps({"error": "Invalid ID"}).encode('utf-8'))
                return
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length).decode('utf-8')
            try:
                data = json.loads(body)
                if not isinstance(data, dict):
                    raise ValueError
            except Exception:
                self._set_json(400)
                self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
                return
            if "price" in data and data.get("source") != "import":
                logger.info(
                    "price field is read-only; ignoring update for product %s (source=%s)",
                    pid,
                    data.get("source"),
                )
                data.pop("price", None)
            conn = ensure_db()
            database.update_product(conn, pid, **data)
            product = row_to_dict(database.get_product(conn, pid))
            if product:
                try:
                    extra_dict = json.loads(rget(product, "extra") or "{}")
                except Exception:
                    extra_dict = {}
                desire_db = rget(product, "desire")
                if desire_db in (None, ""):
                    desire_db = _ensure_desire(product, extra_dict)
                product["desire"] = (desire_db or "").strip() or None
                self._set_json()
                self.wfile.write(json.dumps(product).encode('utf-8'))
            else:
                self.send_error(404)
            return
        if path == "/settings/winner-score":
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length).decode('utf-8')
            try:
                data = json.loads(body)
                if not isinstance(data, dict):
                    raise ValueError
            except Exception:
                self._set_json(400)
                self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
                return
            cfg = config.load_config()
            weights_cfg = cfg.get('weights', {})
            for k, v in data.items():
                try:
                    weights_cfg[k] = float(v)
                except Exception:
                    continue
            cfg['weights'] = weights_cfg
            config.save_config(cfg)
            self._set_json()
            self.wfile.write(json.dumps({"status": "ok"}).encode('utf-8'))
            return
        self.send_error(404)

    def do_PATCH(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path.startswith("/api/products/"):
            try:
                pid = int(path.split("/")[-1])
            except Exception:
                self._set_json(400)
                self.wfile.write(json.dumps({"error": "Invalid ID"}).encode("utf-8"))
                return
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length).decode('utf-8') if length else ""
            try:
                data = json.loads(body or "{}")
                if not isinstance(data, dict):
                    raise ValueError
            except Exception:
                self._set_json(400)
                self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
                return
            logger.info("PATCH product_id=%s fields=%s", pid, list(data.keys()))
            conn = ensure_db()
            prod = database.get_product(conn, pid)
            if not prod:
                logger.warning("PATCH not found product_id=%s", pid)
                self._set_json(404)
                self.wfile.write(json.dumps({"error": "Not found"}).encode('utf-8'))
                return
            allowed = {"desire", "desire_magnitude", "awareness_level", "competition_level"}
            fields = {k: v for k, v in data.items() if k in allowed}
            if fields:
                database.update_product(conn, pid, **fields)
            product = row_to_dict(database.get_product(conn, pid))
            self._set_json()
            self.wfile.write(json.dumps(product).encode('utf-8'))
            return
        if path == "/api/config/winner-weights":
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length).decode('utf-8')
            try:
                data = json.loads(body or "{}")
                if not isinstance(data, dict):
                    raise ValueError
            except Exception:
                self._set_json(400)
                self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
                return
            if data.get("reset"):
                cfg = _apply_weights_reset()
                payload = _build_weights_payload(cfg)
                logger.info("RESET applied weights_order=%s", cfg.get("weights_order"))
                self._set_json()
                self.wfile.write(json.dumps(payload).encode('utf-8'))
                return
              
            weights_in = (
                data.get("weights")
                or data.get("winner_weights")
                or {k: v for k, v in data.items() if k in winner_calc.ALLOWED_FIELDS}
            )
            if not isinstance(weights_in, dict):
                weights_in = {}

            cfg = config.load_config()
            sanitized = _sanitize_weights_map(cfg.get("winner_weights"))
            for key, value in weights_in.items():
                if key in winner_calc.ALLOWED_FIELDS:
                    sanitized[key] = _clamp_weight_value(value)
            for key in DEFAULT_ORDER_LIST:
                sanitized.setdefault(key, 50)

            order_in = data.get("order") or data.get("weights_order")
            if not isinstance(order_in, list) or not order_in:
                order_in = cfg.get("winner_order") or list(DEFAULT_ORDER_LIST)
            order = _normalize_order_list(order_in, sanitized)

            weights_enabled_in = data.get("weights_enabled")
            if isinstance(weights_enabled_in, dict):
                weights_enabled = _sanitize_enabled_map(weights_enabled_in, order)
            else:
                weights_enabled = _sanitize_enabled_map(cfg.get("weights_enabled"), order)

            cfg["winner_weights"] = sanitized
            cfg["winner_order"] = order[:]
            cfg["weights_order"] = order[:]
            cfg["weights_enabled"] = weights_enabled
            cfg["weightsUpdatedAt"] = int(time.time())
            config.save_config(cfg)
            winner_calc.invalidate_weights_cache()

            resp_cfg = dict(cfg)
            resp_cfg.pop("api_key", None)
            resp_cfg["weights"] = dict(cfg["winner_weights"])
            resp_cfg["order"] = list(order)
            resp_cfg["weights_order"] = list(order)

            publish_progress({
                "event": "weights",
                "action": "updated",
                "payload": resp_cfg,
            })
            self._set_json()
            self.wfile.write(json.dumps(resp_cfg).encode('utf-8'))
            return
        self.send_error(404)

    def handle_export_kalodata_minimal(self):
        from .routes_export_minimal import export_kalodata_minimal

        try:
            export_kalodata_minimal(self, ensure_db)
        except Exception:
            logger.exception("Error exportando kalodata minimal")
            self.send_json({"error": "internal_error"}, 500)

    def handle_analyze_titles(self):
        """Endpoint for Title Analyzer.

        Accepts either JSON array of objects or a CSV/XLSX file upload under
        multipart/form-data.  Each item must include a ``title`` field and may
        optionally include ``price`` and ``rating``.  Returns a JSON response
        with the normalized items and placeholder analysis results.
        """
        ctype = self.headers.get('Content-Type', '')
        items = []
        if ctype.startswith('application/json'):
            length = int(self.headers.get('Content-Length', 0))
            raw = self.rfile.read(length)
            try:
                data = json.loads(raw.decode('utf-8'))
                if isinstance(data, list):
                    for obj in data:
                        title = (obj.get('title') or obj.get('name') or '').strip()
                        if not title:
                            continue
                        item = {'title': title}
                        if obj.get('price') is not None:
                            item['price'] = obj.get('price')
                        if obj.get('rating') is not None:
                            item['rating'] = obj.get('rating')
                        items.append(item)
                else:
                    raise ValueError('Expected list')
            except Exception:
                self.send_error(400, 'Invalid JSON')
                return
        elif ctype.startswith('multipart/form-data'):
            filename, data = self._parse_multipart_file()
            if not filename or data is None:
                self.send_error(400, 'No file provided')
                return
            filename = Path(filename).name
            ext = Path(filename).suffix.lower()

            def find_key(keys, patterns):
                for k in keys:
                    sanitized = ''.join(ch.lower() for ch in k if ch.isalnum())
                    for p in patterns:
                        if p in sanitized:
                            return k
                return None

            if ext == '.csv':
                import csv
                text = data.decode('utf-8', errors='ignore')
                reader = csv.DictReader(text.splitlines())
                headers = reader.fieldnames or []
                title_col = find_key(headers, ['title', 'name', 'productname', 'product_name'])
                price_col = find_key(headers, ['price'])
                rating_col = find_key(headers, ['rating', 'stars'])
                for row in reader:
                    title = (row.get(title_col) or '').strip() if title_col else ''
                    if not title:
                        continue
                    item = {'title': title}
                    if price_col and row.get(price_col):
                        try:
                            item['price'] = float(str(row[price_col]).replace(',', '.'))
                        except Exception:
                            pass
                    if rating_col and row.get(rating_col):
                        try:
                            item['rating'] = float(str(row[rating_col]).replace(',', '.'))
                        except Exception:
                            pass
                    items.append(item)
            elif ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
                try:
                    import openpyxl
                except Exception:
                    self.send_error(500, 'openpyxl is required for XLSX files')
                    return
                wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
                ws = wb.active
                rows = ws.iter_rows(values_only=True)
                try:
                    headers = [str(h).strip() if h else '' for h in next(rows)]
                except StopIteration:
                    headers = []
                title_col = find_key(headers, ['title', 'name', 'productname', 'product_name'])
                price_col = find_key(headers, ['price'])
                rating_col = find_key(headers, ['rating', 'stars'])
                title_idx = headers.index(title_col) if title_col in headers else None
                price_idx = headers.index(price_col) if price_col in headers else None
                rating_idx = headers.index(rating_col) if rating_col in headers else None
                for row in rows:
                    if title_idx is None or title_idx >= len(row):
                        continue
                    title = (str(row[title_idx]).strip() if row[title_idx] else '')
                    if not title:
                        continue
                    item = {'title': title}
                    if price_idx is not None and price_idx < len(row) and row[price_idx] is not None:
                        try:
                            item['price'] = float(str(row[price_idx]).replace(',', '.'))
                        except Exception:
                            pass
                    if rating_idx is not None and rating_idx < len(row) and row[rating_idx] is not None:
                        try:
                            item['rating'] = float(str(row[rating_idx]).replace(',', '.'))
                        except Exception:
                            pass
                    items.append(item)
            else:
                self.send_error(400, 'Unsupported file type')
                return
        else:
            self.send_error(400, 'Unsupported Content-Type')
            return

        result = title_analyzer.analyze_titles(items)
        resp = json.dumps({'items': result}).encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.end_headers()
        self.wfile.write(resp)


    def handle_upload(self):
        filename, data = self._parse_multipart_file()
        if not filename or data is None:
            self.send_error(400, "No file provided")
            return
        filename = Path(filename).name
        ext = Path(filename).suffix.lower()
        if ext in (".xlsx", ".xls"):
            tmp_dir = APP_DIR / "uploads"
            tmp_dir.mkdir(exist_ok=True)
            tmp_path = tmp_dir / f"import_{int(time.time()*1000)}{ext}"
            with open(tmp_path, "wb") as f:
                f.write(data)
            conn = ensure_db()
            job_id = database.create_import_job(conn, str(tmp_path))
            threading.Thread(target=_process_import_job, args=(job_id, tmp_path, filename), daemon=True).start()
            self.safe_write(lambda: self.send_json({"task_id": job_id}, status=202))
            return

        if ext == ".csv":
            conn = ensure_db()
            job_config = {"filename": filename, "batch_size": DEFAULT_BATCH_SIZE}
            job_id = database.create_import_job(
                conn,
                status="running",
                phase="parse",
                total=0,
                processed=0,
                config=job_config,
            )
            task_id = str(job_id)
            _update_import_status(
                task_id,
                job_id=job_id,
                state="queued",
                stage="queued",
                done=0,
                total=0,
                error=None,
                imported=0,
                filename=filename,
            )
            _set_import_progress(task_id, pct=0, message="En cola", state="queued")
            csv_bytes = data

            def run_csv():
                _update_import_status(
                    task_id,
                    job_id=job_id,
                    state="running",
                    stage="running",
                    started_at=time.time(),
                )
                _set_import_progress(task_id, pct=5, message="Preparando importación")
                try:
                    def cb(**kwargs):
                        stage = kwargs.get("stage")
                        done = int(kwargs.get("done", 0) or 0)
                        total = int(kwargs.get("total", 0) or 0)
                        extra = {k: v for k, v in kwargs.items() if k not in {"stage", "done", "total"}}
                        if stage == "prepare":
                            _set_import_progress(
                                task_id,
                                pct=8,
                                message="Analizando archivo",
                                done=done,
                                total=total,
                                **extra,
                            )
                        elif stage == "insert":
                            frac = done / max(total, 1) if total else 0.0
                            pct = 20 + min(60, int(round(60 * frac)))
                            msg = f"Insertando registros ({done}/{total})" if total else "Insertando registros"
                            _set_import_progress(
                                task_id,
                                pct=pct,
                                message=msg,
                                done=done,
                                total=total,
                                **extra,
                            )
                        elif stage == "commit":
                            _set_import_progress(
                                task_id,
                                pct=90,
                                message="Guardando cambios",
                                done=done,
                                total=total,
                                **extra,
                            )
                        else:
                            _update_import_status(task_id, **kwargs)

                    imported_count = fast_import(
                        csv_bytes,
                        job_id=job_id,
                        status_cb=cb,
                        source=filename,
                    )
                    job_row = database.get_import_job(conn, job_id)
                    snapshot = _job_payload_from_row(job_row) or {}
                    done_val = int(snapshot.get("processed") or imported_count or 0)
                    total_val = int(snapshot.get("total") or done_val)
                    imported_val = int(snapshot.get("rows_imported") or imported_count or done_val)
                    inserted_ids = database.get_job_product_ids(conn, job_id)
                    rows_imported = len(inserted_ids) or imported_val
                    next_phase = (
                        "enrich"
                        if rows_imported and config.is_auto_fill_ia_on_import_enabled()
                        else "winner"
                    )
                    database.update_import_job_progress(
                        conn,
                        job_id,
                        phase=next_phase,
                        status="running",
                        processed=done_val,
                        total=total_val,
                        rows_imported=rows_imported,
                    )
                    _set_import_progress(
                        task_id,
                        pct=90,
                        message="Importación completada",
                        state="done",
                        stage="done",
                        done=done_val,
                        total=total_val,
                        imported=rows_imported,
                        phase=next_phase,
                    )
                    _schedule_post_import_tasks(job_id, inserted_ids, rows_imported, task_id)
                except Exception as exc:
                    logger.exception("Fast CSV import failed: filename=%s", filename)
                    _update_import_status(
                        task_id,
                        job_id=job_id,
                        state="error",
                        stage="error",
                        error=str(exc),
                        finished_at=time.time(),
                        pct=100,
                        message=f"Error: {exc}",
                    )

            threading.Thread(target=run_csv, daemon=True).start()
            self.safe_write(lambda: self.send_json({"task_id": task_id, "job_id": job_id}, status=202))
            return

        if ext == ".json":
            try:
                payload = json.loads(data.decode("utf-8", errors="ignore"))
            except Exception as exc:
                self.safe_write(lambda: self.send_json({"error": "invalid_json", "detail": str(exc)}, status=400))
                return
            if not isinstance(payload, list):
                self.safe_write(lambda: self.send_json({"error": "invalid_json"}, status=400))
                return
            records = [item for item in payload if isinstance(item, dict)]
            conn = ensure_db()
            total_records = len(records)
            job_config = {
                "filename": filename,
                "batch_size": DEFAULT_BATCH_SIZE,
                "expected": total_records,
            }
            job_id = database.create_import_job(
                conn,
                status="running",
                phase="parse",
                total=total_records,
                processed=0,
                config=job_config,
            )
            task_id = str(job_id)
            _update_import_status(
                task_id,
                job_id=job_id,
                state="queued",
                stage="queued",
                done=0,
                total=total_records,
                error=None,
                imported=0,
                filename=filename,
            )
            _set_import_progress(task_id, pct=0, message="En cola", state="queued")

            def run_json():
                _update_import_status(
                    task_id,
                    job_id=job_id,
                    state="running",
                    stage="running",
                    started_at=time.time(),
                )
                _set_import_progress(task_id, pct=5, message="Preparando importación", total=total_records)
                try:
                    def cb(**kwargs):
                        stage = kwargs.get("stage")
                        done = int(kwargs.get("done", 0) or 0)
                        total = int(kwargs.get("total", total_records) or total_records)
                        extra = {k: v for k, v in kwargs.items() if k not in {"stage", "done", "total"}}
                        if stage == "prepare":
                            _set_import_progress(
                                task_id,
                                pct=8,
                                message="Analizando archivo",
                                done=done,
                                total=total,
                                **extra,
                            )
                        elif stage == "insert":
                            frac = done / max(total, 1) if total else 0.0
                            pct = 20 + min(60, 60 * frac)
                            msg = f"Insertando registros ({done}/{total})" if total else "Insertando registros"
                            _set_import_progress(
                                task_id,
                                pct=pct,
                                message=msg,
                                done=done,
                                total=total,
                                **extra,
                            )
                        elif stage == "commit":
                            _set_import_progress(
                                task_id,
                                pct=82,
                                message="Guardando cambios",
                                done=done,
                                total=total,
                                **extra,
                            )
                        else:
                            _update_import_status(task_id, **kwargs)

                    imported_count = fast_import_records(
                        records,
                        job_id=job_id,
                        status_cb=cb,
                        source=filename,
                    )
                    job_row = database.get_import_job(conn, job_id)
                    snapshot = _job_payload_from_row(job_row) or {}
                    done_val = int(snapshot.get("processed") or imported_count or total_records)
                    total_val = int(snapshot.get("total") or total_records or done_val)
                    imported_val = int(snapshot.get("rows_imported") or imported_count or done_val)
                    inserted_ids = database.get_job_product_ids(conn, job_id)
                    rows_imported = len(inserted_ids) or imported_val
                    next_phase = (
                        "enrich"
                        if rows_imported and config.is_auto_fill_ia_on_import_enabled()
                        else "winner"
                    )
                    database.update_import_job_progress(
                        conn,
                        job_id,
                        phase=next_phase,
                        status="running",
                        processed=done_val,
                        total=total_val,
                        rows_imported=rows_imported,
                    )
                    _set_import_progress(
                        task_id,
                        pct=90,
                        message="Importación completada",
                        state="done",
                        stage="done",
                        done=done_val,
                        total=total_val,
                        imported=rows_imported,
                        phase=next_phase,
                    )
                    _schedule_post_import_tasks(job_id, inserted_ids, rows_imported, task_id)
                except Exception as exc:
                    logger.exception("Fast JSON import failed: filename=%s", filename)
                    _update_import_status(
                        task_id,
                        job_id=job_id,
                        state="error",
                        stage="error",
                        error=str(exc),
                        finished_at=time.time(),
                        pct=100,
                        message=f"Error: {exc}",
                    )

            threading.Thread(target=run_json, daemon=True).start()
            self.safe_write(lambda: self.send_json({"task_id": task_id, "job_id": job_id}, status=202))
            return

        self.safe_write(lambda: self.send_json({"error": "unsupported_format"}, status=400))

    def handle_evaluate_all(self):
        conn = ensure_db()
        api_key = config.get_api_key() or os.environ.get('OPENAI_API_KEY')
        model = config.get_model()
        evaluated = 0
        weights_map = config.get_weights()
        for p_row in database.list_products(conn):
            p = row_to_dict(p_row)
            pid = rget(p, 'id')
            if database.get_scores_for_product(conn, pid):
                continue
            if not (api_key and model):
                continue
            try:
                try:
                    extra = json.loads(rget(p, "extra") or "{}")
                except Exception:
                    extra = {}
                resp = gpt.evaluate_winner_score(
                    api_key,
                    model,
                    {
                        "title": rget(p, "name"),
                        "description": rget(p, "description"),
                        "category": rget(p, "category"),
                        "metrics": extra,
                    },
                )
                scores = resp.get("scores", {})
                justifs = resp.get("justifications", {})
                weighted = sum(
                    scores.get(f, 3) * weights_map.get(f, 0.0)
                    for f in WINNER_SCORE_FIELDS
                )
                raw_score = weighted * 8.0
                pct = ((raw_score - 8.0) / 32.0) * 100.0
                pct = max(0, min(100, round(pct)))
                breakdown = {
                    "scores": scores,
                    "justifications": justifs,
                    "weights": weights_map,
                }
                database.insert_score(
                    conn,
                    product_id=pid,
                    model=model,
                    total_score=0,
                    momentum=0,
                    saturation=0,
                    differentiation=0,
                    social_proof=0,
                    margin=0,
                    logistics=0,
                    summary="",
                    explanations={},
                    winner_score_raw=raw_score,
                    winner_score=pct,
                    winner_score_breakdown=breakdown,
                )
                evaluated += 1
            except Exception:
                continue
            self._set_json()
            self.wfile.write(json.dumps({"evaluated": evaluated}).encode('utf-8'))
            return
        # Legacy evaluation removed; always use Winner Score above

    def handle_setconfig(self):
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length).decode('utf-8')
        try:
            data = json.loads(body)
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
            return
        cfg = config.load_config()
        if 'api_key' in data:
            key = str(data.get('api_key', '')).strip()
            if not key:
                self._set_json(400)
                self.wfile.write(json.dumps({"error": "empty_api_key"}).encode('utf-8'))
                return
            cfg['api_key'] = key
        if 'model' in data and data['model']:
            cfg['model'] = data['model']
        if 'weights' in data and isinstance(data['weights'], dict):
            cfg['weights'] = winner_calc.sanitize_weights(data['weights'])
        if 'autoFillIAOnImport' in data:
            cfg['autoFillIAOnImport'] = bool(data['autoFillIAOnImport'])
        if 'oldness_preference' in data:
            pref = str(data.get('oldness_preference', '')).strip().lower()
            if pref in ("older", "newer"):
                cfg['oldness_preference'] = pref
        config.save_config(cfg)
        self._set_json()
        self.wfile.write(json.dumps({"status": "ok"}).encode('utf-8'))

    def handle_custom_gpt(self):
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length).decode('utf-8')
        try:
            data = json.loads(body)
            prompt = data['prompt']
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid request"}).encode('utf-8'))
            return
        api_key = config.get_api_key() or os.environ.get('OPENAI_API_KEY')
        model = config.get_model()
        if not api_key:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "No API key configured"}).encode('utf-8'))
            return
        try:
            resp = gpt.call_openai_chat(api_key, model, [
                {"role": "system", "content": "Eres un asistente útil."},
                {"role": "user", "content": prompt},
            ])
            content = resp['choices'][0]['message']['content']
            self._set_json()
            self.wfile.write(json.dumps({"response": content}).encode('utf-8'))
        except Exception as exc:
            self._set_json(500)
            self.wfile.write(json.dumps({"error": str(exc)}).encode('utf-8'))

    def handle_prompt_task(self, task: str):
        try:
            canonical = normalize_task(task)
        except KeyError:
            self._set_json(404)
            self.wfile.write(json.dumps({"error": "unknown_task"}).encode('utf-8'))
            return

        length = int(self.headers.get('Content-Length', 0))
        raw_body = self.rfile.read(length) if length else b""
        if raw_body:
            try:
                payload = json.loads(raw_body.decode('utf-8') or "{}")
            except Exception:
                self._set_json(400)
                self.wfile.write(json.dumps({"error": "invalid_json"}).encode('utf-8'))
                return
            if not isinstance(payload, dict):
                self._set_json(400)
                self.wfile.write(json.dumps({"error": "invalid_payload"}).encode('utf-8'))
                return
        else:
            payload = {}

        context_json = payload.get("context_json")
        aggregates = payload.get("aggregates")
        data = payload.get("data")

        try:
            result = gpt.call_gpt(
                canonical,
                context_json=context_json,
                aggregates=aggregates,
                data=data,
            )
        except gpt.InvalidJSONError as exc:
            self._set_json(422)
            self.wfile.write(
                json.dumps({"error": "invalid_model_output", "detail": str(exc)}).encode('utf-8')
            )
            return
        except gpt.OpenAIError as exc:
            self._set_json(502)
            self.wfile.write(json.dumps({"error": "openai_error", "detail": str(exc)}).encode('utf-8'))
            return
        except ValueError as exc:
            self._set_json(404)
            self.wfile.write(json.dumps({"error": "unknown_task", "detail": str(exc)}).encode('utf-8'))
            return
        except Exception as exc:
            logger.exception("Error inesperado en /api/gpt/%s", canonical)
            self._set_json(500)
            self.wfile.write(json.dumps({"error": "internal_error", "detail": str(exc)}).encode('utf-8'))
            return

        self._set_json()
        self.wfile.write(json.dumps(result, ensure_ascii=False).encode('utf-8'))

    def handle_ba_insights(self):
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length).decode('utf-8')
        try:
            payload = json.loads(body)
            product = payload.get("product")
            model = payload.get("model") or "gpt-4o-mini-2024-07-18"
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
            return
        if not isinstance(product, dict) or not product.get("id"):
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Missing product"}).encode('utf-8'))
            return
        api_key = config.get_api_key() or os.environ.get('OPENAI_API_KEY')
        if not api_key:
            self._set_json(503)
            self.wfile.write(json.dumps({"error": "OpenAI no disponible"}).encode('utf-8'))
            return
        try:
            grid_updates, usage, duration = gpt.generate_ba_insights(api_key, model, product)
            logger.info("/api/ba/insights tokens=%s duration=%.2fs", usage.get('total_tokens'), duration)
            self._set_json()
            self.wfile.write(json.dumps({"grid_updates": grid_updates}).encode('utf-8'))
        except gpt.InvalidJSONError:
            self._set_json(502)
            self.wfile.write(json.dumps({"error": "Respuesta IA no es JSON"}).encode('utf-8'))
        except Exception:
            self._set_json(503)
            self.wfile.write(json.dumps({"error": "OpenAI no disponible"}).encode('utf-8'))

    def handle_ia_batch_columns(self):
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length).decode('utf-8')
        try:
            payload = json.loads(body)
            items = payload.get("items")
            model = payload.get("model") or "gpt-4o-mini-2024-07-18"
            if not isinstance(items, list):
                raise ValueError
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
            return
        api_key = config.get_api_key() or os.environ.get('OPENAI_API_KEY')
        if not api_key:
            self._set_json(503)
            self.wfile.write(json.dumps({"error": "OpenAI no disponible"}).encode('utf-8'))
            return
        try:
            ok, ko, usage, duration = gpt.generate_batch_columns(api_key, model, items)
            logger.info("/api/ia/batch-columns tokens=%s duration=%.2fs", usage.get('total_tokens'), duration)
            self._set_json()
            self.wfile.write(json.dumps({"ok": ok, "ko": ko}).encode('utf-8'))
        except gpt.InvalidJSONError:
            self._set_json(502)
            self.wfile.write(json.dumps({"error": "Respuesta IA no es JSON"}).encode('utf-8'))
        except Exception:
            self._set_json(503)
            self.wfile.write(json.dumps({"error": "OpenAI no disponible"}).encode('utf-8'))

    def handle_auto_weights(self):
        """
        Compute recommended weights based on existing scores.
        The idea is to give higher weight to metrics that are on average low
        (indicating scarcity) and lower weight to metrics that are high on average.
        Returns a mapping of metric -> weight.  If no scores exist, returns
        default weights (1.0 for each).
        """
        conn = ensure_db()
        # gather all scores
        rows = database.list_products(conn)
        metrics = ["momentum", "saturation", "differentiation", "social_proof", "margin", "logistics"]
        sums = {m: 0.0 for m in metrics}
        count = 0
        for p in rows:
            scores = database.get_scores_for_product(conn, p["id"])
            if not scores:
                continue
            s = row_to_dict(scores[0])
            count += 1
            for m in metrics:
                try:
                    val = float(rget(s, m, 0.0))
                except Exception:
                    val = 0.0
                sums[m] += val
        if count == 0:
            weights = {m: 1.0 for m in metrics}
        else:
            avg = {m: sums[m] / count for m in metrics}
            # attempt to call GPT to propose weights
            api_key = config.get_api_key() or os.environ.get('OPENAI_API_KEY')
            model = config.get_model()
            recommended = None
            if api_key and model:
                try:
                    prompt = (
                        f"Eres un asistente experto en análisis de productos. Dadas las medias actuales de las métricas "
                        f"(momentum={avg['momentum']:.2f}, saturación={avg['saturation']:.2f}, "
                        f"diferenciación={avg['differentiation']:.2f}, prueba social={avg['social_proof']:.2f}, "
                        f"margen={avg['margin']:.2f}, logística={avg['logistics']:.2f}), "
                        "propón pesos (de 0 a 2) para cada métrica en formato JSON con las claves exactas: "
                        "momentum, saturation, differentiation, social_proof, margin, logistics."
                    )
                    resp = gpt.call_openai_chat(api_key, model, [
                        {"role": "system", "content": "Eres un asistente experto en scoring de productos."},
                        {"role": "user", "content": prompt},
                    ])
                    content = resp['choices'][0]['message']['content']
                    # Attempt to parse JSON from content
                    try:
                        recommended = json.loads(content)
                    except Exception:
                        recommended = None
                except Exception:
                    recommended = None
            if recommended and all(k in recommended for k in metrics):
                # use recommended weights
                weights = {k: float(recommended.get(k, 1.0)) for k in metrics}
            else:
                # fallback: inverse of average (scarce factors get higher weight)
                weights = {m: (1.0 / (avg[m] + 1e-6)) for m in metrics}
            # normalise so sum of weights equals len(metrics)
            total = sum(weights.values()) or 1.0
            factor = float(len(metrics)) / total
            for k in weights:
                weights[k] *= factor
        self._set_json()
        self.wfile.write(json.dumps(weights).encode('utf-8'))

    def _collect_samples_for_weights(self):
        """Gather products with Winner Score breakdown and success metric."""
        conn = ensure_db()
        rows = database.list_products(conn)
        samples = []
        metric_key = None
        for p in rows:
            try:
                extra = json.loads(p["extra"] or "{}")
            except Exception:
                extra = {}
            success = None
            if metric_key and metric_key in extra:
                try:
                    success = float(extra[metric_key])
                except Exception:
                    success = None
            if success is None:
                for key in ("orders", "revenue", "gmv", "sales", "units"):
                    if key in extra:
                        try:
                            success = float(extra[key])
                            metric_key = metric_key or key
                            break
                        except Exception:
                            continue
            if success is None:
                continue
            scores_rows = database.get_scores_for_product(conn, p["id"])
            if not scores_rows:
                continue
            srow = scores_rows[0]
            try:
                breakdown = json.loads(srow["winner_score_breakdown"] or "{}")
                scores = breakdown.get("scores") or {}
            except Exception:
                continue
            if not scores or any(k not in scores for k in WINNER_SCORE_FIELDS):
                continue
            sample = {k: float(scores[k]) for k in WINNER_SCORE_FIELDS}
            sample[metric_key] = success
            samples.append(sample)
            if len(samples) >= 50:
                break
        return samples, metric_key

    def handle_scoring_v2_auto_weights_gpt(self):
        length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(length).decode("utf-8") if length else ""
        try:
            data = json.loads(body) if body else {}
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
            return
        features = [f for f in (data.get("features") or WINNER_SCORE_FIELDS) if f in winner_calc.ALLOWED_FIELDS]
        samples_in = data.get("data_sample") or []
        target = data.get("target") or ""
        if not samples_in or not target:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Datos insuficientes"}).encode('utf-8'))
            return
        samples = []
        for s in samples_in:
            if "target" not in s:
                continue
            row = {k: float(s.get(k, 0.0)) for k in features}
            row[target] = float(s.get("target", 0.0))
            samples.append(row)
        api_key = config.get_api_key() or os.environ.get('OPENAI_API_KEY')
        model = config.get_model()
        if not api_key or not model:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "No API key configured"}).encode('utf-8'))
            return
        # --- pedir a GPT (o devolverá fallback desde gpt.py) ---
        result = gpt.recommend_winner_weights(api_key, model, samples, target)
        weights_raw = result.get("weights", {}) or {}
        notes = result.get("justification", "")

        # Filtra a campos permitidos + clamp 0..100 enteros
        allowed = list(winner_calc.ALLOWED_FIELDS)
        final_weights = {}
        for k in allowed:
            v = weights_raw.get(k, 0)
            try:
                v = float(v)
            except Exception:
                v = 0.0
            v = max(0.0, min(100.0, v))
            final_weights[k] = int(round(v))

        # ORDER: por peso desc; a igualdad conserva el orden previo
        prev_settings = winner_calc.load_settings()
        prev_order = (
            prev_settings.get("winner_order")
            or prev_settings.get("weights_order")
        )
        if not isinstance(prev_order, list) or not prev_order:
            from .services.config import DEFAULT_ORDER as DEFAULT_WINNER_ORDER

            prev_order = list(DEFAULT_WINNER_ORDER)
        rank = {k: i for i, k in enumerate(prev_order)}
        order = sorted(
            final_weights.keys(),
            key=lambda k: (
                -final_weights[k],
                rank.get(k, len(rank) + allowed.index(k) if k in allowed else len(rank) + 999),
            ),
        )

        int_weights = {k: int(final_weights.get(k, 0)) for k in allowed}
        new_order = list(order)

        cfg = config.load_config()
        cfg["winner_weights"] = int_weights
        cfg["winner_order"] = new_order[:]
        cfg["weights_order"] = new_order[:]
        cfg["weights_enabled"] = _sanitize_enabled_map(
            cfg.get("weights_enabled"), new_order
        )
        cfg["weightsUpdatedAt"] = int(time.time())
        config.save_config(cfg)
        winner_calc.invalidate_weights_cache()

        # Logs (útiles para ti): ahora ai_raw/ints son lo mismo (0..100 independientes)
        logger.info(
            "ai_raw=%s enabled_only=%s ints=%s order=%s sum=%s",
            int_weights,
            int_weights,
            int_weights,
            new_order,
            sum(int_weights.values()),
        )

        # Respuesta para el frontend; los valores ya quedaron persistidos en backend.
        resp = {
            "weights": int_weights,      # 0..100 independientes
            "weights_order": new_order,  # prioridad explícita
            "order": new_order,
            "method": "gpt",
            "diagnostics": {"notes": notes},
        }
        self._set_json()
        self.wfile.write(json.dumps(resp).encode('utf-8'))

    def handle_scoring_v2_auto_weights_stat(self):
        length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(length).decode("utf-8") if length else ""
        try:
            data = json.loads(body) if body else {}
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
            return
        features = [f for f in (data.get("features") or WINNER_SCORE_FIELDS) if f in winner_calc.ALLOWED_FIELDS]
        samples_in = data.get("data_sample") or []
        target = data.get("target") or ""
        if not samples_in or not target or len(samples_in) < 2:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Datos insuficientes"}).encode('utf-8'))
            return
        ys = [float(s.get("target", 0.0)) for s in samples_in]
        mean_y = sum(ys) / len(ys)
        denom_y = math.sqrt(sum((y - mean_y) ** 2 for y in ys)) or 1.0
        weights: Dict[str, float] = {}
        for field in features:
            xs = [float(s.get(field, 0.0)) for s in samples_in]
            mean_x = sum(xs) / len(xs)
            denom_x = math.sqrt(sum((x - mean_x) ** 2 for x in xs)) or 1.0
            num = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys))
            corr = abs(num / (denom_x * denom_y)) if denom_x and denom_y else 0.0
            weights[field] = corr

        # weights como correlación absoluta -> reescala a 0..100 independientes
        weights01 = {k: float(weights.get(k, 0.0)) for k in features}
        maxv = max(weights01.values() or [0.0])
        weights_raw = {
            k: (v / maxv * 100.0 if maxv > 0 else 50.0)
            for k, v in weights01.items()
        }

        allowed = list(winner_calc.ALLOWED_FIELDS)
        final_weights = {}
        for k in allowed:
            v = weights_raw.get(k, 0.0)
            v = max(0.0, min(100.0, float(v)))
            final_weights[k] = int(round(v))

        prev_settings = winner_calc.load_settings()
        prev_order = prev_settings.get("weights_order") or list(allowed)
        order = sorted(final_weights.keys(), key=lambda k: (-final_weights[k], prev_order.index(k) if k in prev_order else 999))

        resp = {
            "weights": final_weights,        # 0..100 independientes
            "weights_order": order,
            "order": order,
            "method": "stat",
            "diagnostics": {"n": len(samples_in)},
        }
        self._set_json()
        self.wfile.write(json.dumps(resp).encode('utf-8'))

    def handle_scoring_v2_gpt_evaluate(self):
        """Endpoint that evaluates Winner Score variables via GPT."""

        length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(length).decode("utf-8") if length else ""
        try:
            data = json.loads(body) if body else {}
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode("utf-8"))
            return

        title = data.get("title") or data.get("name") or ""
        description = data.get("description") or ""
        category = data.get("category") or ""
        metrics = data.get("metrics") or {}

        api_key = config.get_api_key() or os.environ.get("OPENAI_API_KEY")
        model = config.get_model()
        scores, justifs, sources = gpt.compute_numeric_scores(metrics, {})
        need = [f for f in WINNER_SCORE_FIELDS if f not in scores]
        if need:
            if not api_key or not model:
                self._set_json(400)
                self.wfile.write(json.dumps({"error": "No API key configured"}).encode("utf-8"))
                return
            try:
                resp = gpt.evaluate_winner_score(
                    api_key,
                    model,
                    {
                        "title": title,
                        "description": description,
                        "category": category,
                        "metrics": metrics,
                    },
                )
            except Exception as exc:
                self._set_json(500)
                self.wfile.write(json.dumps({"error": str(exc)}).encode("utf-8"))
                return
            rs = resp.get("scores", {})
            js = resp.get("justifications", {})
            for f in need:
                scores[f] = rs.get(f, 3)
                justifs[f] = js.get(f, "")
                sources[f] = "gpt"
        out = {**scores, "justificacion": justifs, "source": sources}
        self._set_json()
        self.wfile.write(json.dumps(out).encode("utf-8"))

    def handle_scoring_v2_gpt_summary(self):
        """Generate an executive summary of top products using GPT."""

        length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(length).decode("utf-8") if length else ""
        try:
            data = json.loads(body) if body else {}
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode("utf-8"))
            return

        products = data.get("products") or []
        if not isinstance(products, list) or not products:
            self._set_json(400)
            self.wfile.write(
                json.dumps({"error": "No products provided"}).encode("utf-8")
            )
            return

        api_key = config.get_api_key() or os.environ.get("OPENAI_API_KEY")
        model = config.get_model()
        if not api_key or not model:
            self._set_json(400)
            self.wfile.write(
                json.dumps({"error": "No API key configured"}).encode("utf-8")
            )
            return

        try:
            summary = gpt.summarize_top_products(api_key, model, products)
        except Exception as exc:
            self._set_json(500)
            self.wfile.write(json.dumps({"error": str(exc)}).encode("utf-8"))
            return

        self._set_json()
        self.wfile.write(json.dumps({"summary": summary}).encode("utf-8"))

    def handle_scoring_v2_generate(self):
        """Compute Winner Score for selected products."""

        parsed = urlparse(self.path)
        params = parse_qs(parsed.query)
        debug = params.get("debug", ["0"])[0] == "1"

        length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(length).decode("utf-8") if length else ""
        try:
            data = json.loads(body) if body else {}
            ids = data.get("product_ids") or data.get("ids") or []
            if ids and not isinstance(ids, list):
                raise ValueError
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode("utf-8"))
            return

        logger.info("Winner Score generate: ids_length=%d", len(ids))
        conn = ensure_db()
        result = winner_calc.generate_winner_scores(conn, product_ids=ids or None, debug=debug)
        resp = {
            "ok": True,
            "processed": result.get("processed", 0),
            "updated": result.get("updated", 0),
            "weights_all": result.get("weights_all"),
            "weights_eff": result.get("weights_eff"),
        }
        if debug:
            resp["diag"] = result.get("diag", {})
        self._set_json()
        self.wfile.write(json.dumps(resp).encode("utf-8"))

    def handle_winner_score_explain(self, parsed):
        """Return detailed Winner Score components for debugging."""

        params = parse_qs(parsed.query)
        ids_param = params.get("ids", [""])[0]
        ids = [int(x) for x in ids_param.split(",") if x.strip()]

        conn = ensure_db()
        weights, order, enabled = winner_calc.load_winner_settings()

        if ids:
            placeholders = ",".join("?" for _ in ids)
            cur = conn.execute(
                f"SELECT * FROM products WHERE id IN ({placeholders})",
                tuple(ids),
            )
            rows = cur.fetchall()
        else:
            rows = database.list_products(conn)

        winner_calc.prepare_oldness_bounds(rows)

        data: Dict[str, Any] = {}
        for row in rows:
            res = winner_calc.compute_winner_score_v2(row, weights, order, enabled)
            sf = res.get("score_float") or 0.0
            score_raw = max(0.0, min(1.0, sf)) * 100.0
            data[row["id"]] = {
                "present": res.get("present_fields", []),
                "missing": res.get("missing_fields", []),
                "effective_weights": {k: round(v, 3) for k, v in res.get("effective_weights", {}).items()},
                "score_raw": score_raw,
                "score_int": int(round(score_raw)),
            }

        self._set_json()
        self.wfile.write(json.dumps(data).encode("utf-8"))

    def handle_create_list(self):
        """Create a new user defined list (group) of products."""
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length).decode('utf-8') if length else ''
        try:
            data = json.loads(body) if body else {}
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
            return
        name = (data.get('name') or '').strip()
        if not name:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Nombre no proporcionado"}).encode('utf-8'))
            return
        conn = ensure_db()
        list_id = database.create_list(conn, name)
        self._set_json()
        self.wfile.write(json.dumps({"id": list_id, "name": name}).encode('utf-8'))

    def handle_delete_list(self):
        """Delete an existing list by ID with options to move or remove products."""
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length).decode('utf-8') if length else ''
        try:
            data = json.loads(body) if body else {}
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
            return
        try:
            lid = int(data.get('id'))
            mode = data.get('mode', 'remove')
            tgt = data.get('targetGroupId')
            if tgt is not None:
                tgt = int(tgt)
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Datos inválidos"}).encode('utf-8'))
            return
        conn = ensure_db()
        try:
            result = database.delete_list(conn, lid, mode=mode, target_list_id=tgt)
            self._set_json()
            self.wfile.write(json.dumps(result).encode('utf-8'))
        except Exception as exc:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": str(exc)}).encode('utf-8'))

    def handle_add_to_list(self):
        """Add one or more products to a list."""
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length).decode('utf-8') if length else ''
        try:
            data = json.loads(body) if body else {}
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
            return
        try:
            lid = int(data.get('id'))
            ids = [int(x) for x in data.get('ids', [])]
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Datos inválidos"}).encode('utf-8'))
            return
        conn = ensure_db()
        for pid in ids:
            database.add_product_to_list(conn, lid, pid)
        self._set_json()
        self.wfile.write(json.dumps({"added": len(ids)}).encode('utf-8'))

    def handle_shutdown(self):
        """Shutdown the HTTP server."""
        self._set_json()
        self.wfile.write(json.dumps({"ok": True}).encode('utf-8'))
        threading.Thread(target=self.server.shutdown, daemon=True).start()

    def handle_delete(self):
        """Delete one or more products specified in the request body.

        Expects a JSON payload with an "ids" array of product IDs to delete.
        Returns a JSON object with the number of deleted records.  If any
        ID is invalid, it will be skipped.
        """
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length).decode('utf-8') if length > 0 else '{}'
        try:
            data = json.loads(body or '{}')
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
            return
        ids = data.get('ids')
        if not isinstance(ids, list):
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Missing or invalid ids"}).encode('utf-8'))
            return
        conn = ensure_db()
        deleted = 0
        deleted_ids: List[int] = []
        for pid in ids:
            try:
                pid_int = int(pid)
            except Exception:
                continue
            try:
                database.delete_product(conn, pid_int)
                deleted += 1
                deleted_ids.append(pid_int)
            except Exception:
                continue
        if deleted_ids:
            publish_progress({
                "event": "delete",
                "ids": deleted_ids,
                "deleted": deleted,
            })
        self._set_json()
        self.wfile.write(json.dumps({"deleted": deleted}).encode('utf-8'))

    def handle_remove_from_list(self):
        """Remove products from a specific list without deleting them globally.

        Expects JSON payload with ``list_id`` and ``ids`` array. Removes each product from the list.
        Returns number of associations removed.
        """
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length).decode('utf-8') if length > 0 else '{}'
        try:
            data = json.loads(body or '{}')
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid JSON"}).encode('utf-8'))
            return
        list_id = data.get('list_id')
        ids = data.get('ids')
        try:
            list_id_int = int(list_id)
        except Exception:
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Invalid list_id"}).encode('utf-8'))
            return
        if not isinstance(ids, list):
            self._set_json(400)
            self.wfile.write(json.dumps({"error": "Missing or invalid ids"}).encode('utf-8'))
            return
        conn = ensure_db()
        removed = 0
        for pid in ids:
            try:
                pid_int = int(pid)
            except Exception:
                continue
            try:
                database.remove_product_from_list(conn, list_id_int, pid_int)
                removed += 1
            except Exception:
                continue
        self._set_json()
        self.wfile.write(json.dumps({"removed": removed}).encode('utf-8'))


def run(host: str = '127.0.0.1', port: int = 8000):
    ensure_db()
    resume_incomplete_imports()
    httpd = HTTPServer((host, port), RequestHandler)
    print(f"Servidor iniciado en http://{host}:{port}")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("Apagando servidor...")
    httpd.server_close()


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description="Web UI for Product Research Copilot")
    parser.add_argument('--host', default='127.0.0.1', help='Host IP to bind')
    parser.add_argument('--port', default=8000, type=int, help='Port number')
    args = parser.parse_args()
    run(args.host, args.port)
