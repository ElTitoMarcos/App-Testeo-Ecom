"""Export selected products to an XLSX workbook for Vega."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
import re

from flask import Blueprint, current_app, request, send_file
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import requests


bp_export = Blueprint("export_bp", __name__)


UA = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
    )
}
IMG_W = 192
ROW_PAD = 22


def _px_to_width(px: int) -> float:
    """Convert pixel width to Excel's column width units."""

    return max(10.0, (px - 5) / 7.0)


def _clean_num(value):
    """Return a float from strings like ``1.2K`` or ``12%``."""

    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:  # pragma: no cover - defensive
            return None
    s = str(value).strip()
    if not s:
        return None
    s = s.replace(",", "")
    mult = 1.0
    match = re.search(r"([\d.]+)\s*([KMB])", s, re.IGNORECASE)
    if match:
        val = float(match.group(1))
        mult = {"K": 1e3, "M": 1e6, "B": 1e9}[match.group(2).upper()]
        return val * mult
    if s.endswith("%"):
        try:
            return float(s[:-1])
        except Exception:
            return None
    s = s.replace("$", "")
    try:
        return float(s)
    except Exception:
        return None


def _download_png_resized(url: str, out_path: Path) -> bool:
    """Download ``url`` and save a PNG resized to ``IMG_W`` pixels wide."""

    try:
        resp = requests.get(url, headers=UA, timeout=12)
        resp.raise_for_status()
        image = PILImage.open(BytesIO(resp.content)).convert("RGBA")
        if image.width != IMG_W:
            ratio = IMG_W / float(image.width or 1)
            image = image.resize(
                (IMG_W, max(1, int(image.height * ratio))), PILImage.LANCZOS
            )
        image.save(out_path, format="PNG")
        return True
    except Exception:
        return False


def _find_source(item: dict) -> str:
    for key in ("product_url", "source_url", "listing_url", "url", "link"):
        val = item.get(key)
        if val:
            return str(val)
    src = item.get("source")
    return str(src) if src else ""


def _fetch_rows(ids, columns):
    provider = current_app.config.get("ROW_PROVIDER")
    return provider(ids, columns) if callable(provider) else []


@bp_export.route("/api/export", methods=["POST"])
def export_xlsx():
    """Create an XLSX workbook with Vega and full product context sheets."""

    payload = request.get_json(force=True) or {}
    ids = payload.get("ids") or []
    columns = payload.get("columns") or []
    rows = _fetch_rows(ids, columns) or []

    workbook = Workbook()
    ws_input = workbook.active
    ws_input.title = "VEGA_INPUT"
    ws_full = workbook.create_sheet("PRODUCTS_FULL")

    header_fill = PatternFill("solid", fgColor="1f2347")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    topwrap = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="333333")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    input_cols = [
        ("Product Name", "name"),
        ("Source", "__source__"),
        ("Desire", "desire"),
        ("Desire Magnitude", "desire_magnitude"),
        ("Awareness Level", "awareness_level"),
        ("Competition Level", "competition_level"),
    ]

    for idx, (title, _key) in enumerate(input_cols, start=1):
        cell = ws_input.cell(row=1, column=idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws_input.column_dimensions[cell.column_letter].width = 28

    full_cols = [
        ("ID", "id"),
        ("Image", "image"),
        ("Product Name", "name"),
        ("Category", "category"),
        ("Price (num)", "price"),
        ("Rating (num)", "rating"),
        ("Units Sold (num)", "units_sold"),
        ("Revenue (num)", "revenue"),
        ("Conversion Rate (%)", "conversion_rate"),
        ("Launch Date", "launch_date"),
        ("Date Range", "date_range"),
        ("Desire", "desire"),
        ("Desire Magnitude", "desire_magnitude"),
        ("Awareness Level", "awareness_level"),
        ("Competition Level", "competition_level"),
        ("Winner Score", "winner_score"),
        ("Source", "__source__"),
    ]

    for idx, (title, _key) in enumerate(full_cols, start=1):
        cell = ws_full.cell(row=1, column=idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws_full.column_dimensions[cell.column_letter].width = 20

    alias_map = {
        "name": ("name", "nombre", "title"),
        "category": ("category", "categoria"),
        "price": ("price", "precio"),
        "rating": ("rating",),
        "units_sold": ("units_sold", "units", "items_sold"),
        "revenue": ("revenue", "ventas", "gmv"),
        "conversion_rate": ("conversion_rate", "conversion"),
        "launch_date": ("launch_date", "launch"),
        "date_range": ("date_range", "date-range"),
        "desire": ("desire", "benefit", "claim"),
        "desire_magnitude": ("desire_magnitude", "magnitud_deseo", "magnitude"),
        "awareness_level": ("awareness_level", "awareness"),
        "competition_level": ("competition_level", "competition"),
        "winner_score": ("winner_score", "winnerScore", "score"),
        "image": ("image", "image_url", "img", "thumbnail"),
    }

    def map_val(item: dict, key: str):
        if key in {"__source__", "image"}:
            return None
        for alias in alias_map.get(key, (key,)):
            if alias in item and item[alias] not in (None, ""):
                return item[alias]
        return None

    buffer = BytesIO()
    with TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)

        for row_idx, item in enumerate(rows, start=2):
            src_value = _find_source(item)
            desire_val = map_val(item, "desire") or ""
            input_data = {
                "name": map_val(item, "name") or "",
                "__source__": src_value,
                "desire": desire_val,
                "desire_magnitude": map_val(item, "desire_magnitude") or "",
                "awareness_level": map_val(item, "awareness_level") or "",
                "competition_level": map_val(item, "competition_level") or "",
            }
            for col_idx, (_title, key) in enumerate(input_cols, start=1):
                cell = ws_input.cell(row=row_idx, column=col_idx, value=input_data.get(key, ""))
                cell.alignment = topwrap
                cell.border = border
        ws_input.freeze_panes = "A2"

        image_col_index = next(idx for idx, (_title, key) in enumerate(full_cols, start=1) if key == "image")
        image_col_letter = get_column_letter(image_col_index)

        number_formats = {
            "price": "0.00",
            "rating": "0.00",
            "units_sold": "#,##0",
            "revenue": "0.00",
            "conversion_rate": "0.00",
            "winner_score": "0",
        }

        for row_idx, item in enumerate(rows, start=2):
            src_value = _find_source(item)
            full_values = {
                "id": item.get("id"),
                "image": item.get("image")
                or item.get("image_url")
                or item.get("img")
                or item.get("thumbnail"),
                "name": map_val(item, "name"),
                "category": map_val(item, "category"),
                "price": _clean_num(map_val(item, "price")),
                "rating": _clean_num(map_val(item, "rating")),
                "units_sold": _clean_num(map_val(item, "units_sold")),
                "revenue": _clean_num(map_val(item, "revenue")),
                "conversion_rate": _clean_num(map_val(item, "conversion_rate")),
                "launch_date": map_val(item, "launch_date"),
                "date_range": map_val(item, "date_range"),
                "desire": map_val(item, "desire"),
                "desire_magnitude": map_val(item, "desire_magnitude"),
                "awareness_level": map_val(item, "awareness_level"),
                "competition_level": map_val(item, "competition_level"),
                "winner_score": _clean_num(map_val(item, "winner_score")),
                "__source__": src_value,
            }

            for col_idx, (_title, key) in enumerate(full_cols, start=1):
                if key == "image":
                    ws_full.cell(row=row_idx, column=col_idx, value=None).alignment = topwrap
                    continue
                value = full_values.get(key)
                cell = ws_full.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = topwrap
                cell.border = border
                fmt = number_formats.get(key)
                if fmt and value is not None:
                    cell.number_format = fmt

            image_url = full_values.get("image")
            if image_url and isinstance(image_url, str) and image_url.startswith("http"):
                tmp_file = tmp_path / f"image_{row_idx}.png"
                if _download_png_resized(image_url, tmp_file):
                    xl_image = XLImage(str(tmp_file))
                    xl_image.width = IMG_W
                    xl_image.height = IMG_W
                    anchor = f"{image_col_letter}{row_idx}"
                    try:
                        ws_full.add_image(xl_image, anchor)
                    except TypeError:
                        xl_image.anchor = anchor
                        ws_full.add_image(xl_image)
                    ws_full.column_dimensions[image_col_letter].width = _px_to_width(IMG_W + 20)
                    ws_full.row_dimensions[row_idx].height = IMG_W + ROW_PAD

        ws_full.freeze_panes = "A2"

        workbook._sheets = [ws_input, ws_full]
        workbook.save(buffer)
    buffer.seek(0)

    filename = f"export_vega_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

