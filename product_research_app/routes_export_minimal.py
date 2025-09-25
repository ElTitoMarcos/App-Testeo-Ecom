"""Utilities to expose the Kalodata minimal XLSX export endpoint."""

from __future__ import annotations

import bisect
import io
import json
import logging
import math
import os
import re
import threading
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Mapping, MutableMapping, Optional, Sequence, Tuple

import pandas as pd
import requests
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.cell_range import CellRange

from . import database
from .utils import sanitize_product_name
from .utils.db import row_to_dict

logger = logging.getLogger(__name__)

APP_DIR = Path(__file__).resolve().parent
STATE_DIR = (APP_DIR / ".." / "state").resolve()
EXPORT_DIR = (APP_DIR / ".." / "exports").resolve()

_SEQ_LOCK = threading.Lock()
_SEQ_FILE_NAME = "export_seq.json"
_DESIRE_MAG_HEADERS = {
    "desire magnitude",
    "desire_magnitude",
    "desire mag",
    "desire_mag",
    "magnitud del deseo",
}


def _seq_path() -> Path:
    return STATE_DIR / _SEQ_FILE_NAME


def next_export_id() -> int:
    seq_path = _seq_path()
    with _SEQ_LOCK:
        try:
            STATE_DIR.mkdir(parents=True, exist_ok=True)
        except Exception:
            logger.exception("failed to ensure export state directory exists: %s", STATE_DIR)
        current = 0
        if seq_path.exists():
            try:
                with seq_path.open("r", encoding="utf-8") as fh:
                    payload = json.load(fh) or {}
                current = int(payload.get("export_seq", 0))
            except Exception:
                logger.exception("failed to read export sequence from %s", seq_path)
                current = 0
        current += 1
        tmp_path = seq_path.with_suffix(f"{seq_path.suffix}.tmp")
        try:
            with tmp_path.open("w", encoding="utf-8") as fh:
                json.dump({"export_seq": current}, fh)
            os.replace(tmp_path, seq_path)
        except Exception:
            logger.exception("failed to persist export sequence to %s", seq_path)
        return current


def build_export_filename(base_dir: os.PathLike[str] | str) -> Path:
    base_path = Path(base_dir)
    try:
        base_path.mkdir(parents=True, exist_ok=True)
    except Exception:
        logger.exception("failed to ensure export directory exists: %s", base_path)
    export_id = next_export_id()
    return base_path / f"Analisis_Export_{export_id:04d}.xlsx"


def _is_desire_magnitude_header(value: Any) -> bool:
    return isinstance(value, str) and value.strip().lower() in _DESIRE_MAG_HEADERS


ResponseHandler = Any
EnsureDbFn = Callable[[], Any]

BASE_COLUMNS: Sequence[str] = (
    "Product Name",
    "TikTokUrl",
    "KalodataUrl",
    "Img_url",
    "Image Preview",
    "Category",
    "Price($)",
    "Product Rating",
    "Item Sold",
    "Total Revenue($)",
    "Launch Date",
    "Desire",
    "Desire Magnitude",
    "Awareness Level",
    "Competition Level",
)

COLUMNS: Sequence[str] = (
    "image preview",
    "name",
    "category",
    "launch date",
    "price",
    "rating",
    "revenue",
    "items sold",
    "desire",
    "desire magnitude",
    "awareness level",
    "competition level",
    "TikTokUrl",
    "KalodataUrl",
    "Img_url",
)


def reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    desired_labels = [
        "image preview",
        "name",
        "category",
        "launch date",
        "price",
        "rating",
        "revenue",
        "items sold",
        "desire",
        "desire magnitude",
        "awareness level",
        "competition level",
    ]

    lower_map = {c.lower(): c for c in df.columns}

    synonyms = {
        "image preview": [
            "image preview",
            "image_preview",
            "preview image",
            "thumbnail",
            "miniatura",
            "img",
            "product image",
            "image url",
            "image_url",
            "img_url",
            "imagen",
            "image",
        ],
        "name": ["name", "title", "product name", "nombre", "producto"],
        "category": ["category", "categories", "categoria", "categoría"],
        "launch date": [
            "launch date",
            "launch_date",
            "launched at",
            "launched_at",
            "release date",
            "release_date",
            "first seen",
            "first_seen",
            "fecha de lanzamiento",
            "fecha lanzamiento",
            "date launched",
            "launchdate",
        ],
        "price": ["price", "precio", "price($)", "price usd", "price$"],
        "rating": ["rating", "rate", "stars", "product rating", "valoración", "valoracion"],
        "revenue": [
            "revenue",
            "total revenue",
            "total revenue($)",
            "income",
            "ingresos",
        ],
        "items sold": [
            "items sold",
            "units_sold",
            "units sold",
            "item sold",
            "sold",
            "unidades vendidas",
        ],
        "desire": ["desire", "desires"],
        "desire magnitude": [
            "desire magnitude",
            "desire_magnitude",
            "desire mag",
            "desire_mag",
            "magnitud del deseo",
        ],
        "awareness level": [
            "awareness level",
            "awareness",
            "nivel de awareness",
            "nivel awareness",
        ],
        "competition level": [
            "competition level",
            "competition",
            "nivel de competencia",
        ],
    }

    resolved: List[str] = []
    rename_map: Dict[str, str] = {}
    for canonical in desired_labels:
        found = None
        for variant in synonyms.get(canonical, [canonical]):
            key = variant.lower()
            if key in lower_map:
                found = lower_map[key]
                break
        if found and found not in resolved:
            resolved.append(found)
            if found != canonical:
                rename_map[found] = canonical

    remaining = [c for c in df.columns if c not in resolved]

    ordered_cols = resolved + remaining
    df_out = df.reindex(columns=ordered_cols)

    if rename_map:
        df_out = df_out.rename(columns=rename_map)

    cols_lower = {c.lower(): c for c in df_out.columns}
    if "image preview" in cols_lower and "image" in cols_lower:
        df_out = df_out.drop(columns=[cols_lower["image"]])

    return df_out

TEXT_FIELD_KEYS: Dict[str, Sequence[str]] = {
    "Product Name": ("name", "product_name", "title"),
    "TikTokUrl": (
        "tiktok_url",
        "TikTokUrl",
        "tiktokurl",
        "tiktok",
        "tiktok link",
        "TikTok URL",
        "TikTok Url",
    ),
    "KalodataUrl": ("kalodata_url", "KalodataUrl", "kalodataurl", "Kalodata URL"),
    "Img_url": (
        "image_url",
        "image",
        "img_url",
        "img",
        "imageurl",
        "imagelink",
        "Img_url",
        "Img Url",
    ),
    "Category": ("category", "category_path", "Category"),
    "Launch Date": ("launch_date", "Launch Date", "launchdate"),
}

NUMERIC_FIELD_KEYS: Dict[str, Sequence[str]] = {
    "Price($)": ("price", "Price($)", "price$", "price_usd"),
    "Product Rating": ("rating", "Product Rating", "productrating"),
    "Item Sold": ("units_sold", "items_sold", "Item Sold", "sold"),
    "Revenue($)": ("revenue", "Revenue($)", "total_revenue"),
    "Live Revenue($)": ("live_revenue", "Live Revenue($)", "live revenue"),
    "Video Revenue($)": ("video_revenue", "Video Revenue($)", "video revenue"),
    "Shopping Mall Revenue($)": (
        "shopping_mall_revenue",
        "Shopping Mall Revenue($)",
        "shopping mall revenue",
    ),
}

DESIRE_TEXT_KEYS: Sequence[str] = (
    "desire_statement",
    "desire",
    "desires",
    "customer_desire",
    "desire_text",
    "ai_desire_label",
)
DESIRE_MAGNITUDE_KEYS: Sequence[str] = (
    "desire_magnitude",
    "desiremag",
    "desire_mag",
)

DESIRE_SCORE_KEYS: Sequence[str] = (
    "desire_score",
    "_desire_score",
    "ai_desire_score",
)

AWARENESS_KEYS: Sequence[str] = (
    "awareness_level_label",
    "awareness_level",
    "awareness",
    "awareness_score",
)

COMPETITION_KEYS: Sequence[str] = (
    "competition_level_label",
    "competition_level",
    "competition",
    "competition_score",
)

_NUMERIC_STRIP_RE = re.compile(r"[\s\$,€£%]")

_AWARENESS_LABELS = {
    0: "Unaware",
    1: "Problem-aware",
    2: "Solution-aware",
    3: "Product-aware",
    4: "Most aware",
}

_IMG_MAX_WIDTH = 240
_IMG_MAX_HEIGHT = 160
_IMG_ROW_HEIGHT = 150

_IMAGE_SESSION = requests.Session()
_IMAGE_SESSION.headers.update({"User-Agent": "product-research-exporter/1.0"})
_IMG_CACHE: Dict[str, Optional[bytes]] = {}


def _build_header_index_map(ws) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for cell in ws[1]:
        value = cell.value
        if isinstance(value, str):
            mapping[value.strip().lower()] = cell.col_idx
    return mapping


def _resolve_column_index(mapping: Mapping[str, int], *names: str) -> Optional[int]:
    for name in names:
        key = name.strip().lower()
        if key in mapping:
            return mapping[key]
    return None


def export_kalodata_minimal(handler: ResponseHandler, ensure_db: EnsureDbFn) -> None:
    start = time.perf_counter()
    length = _read_length(handler.headers)
    raw_body = handler.rfile.read(length) if length else b""
    try:
        payload = json.loads(raw_body.decode("utf-8") or "{}")
    except Exception:
        handler.send_json({"error": "invalid_json"}, 400)
        return

    ids = payload.get("ids")
    if not isinstance(ids, list) or not ids:
        handler.send_json({"error": "ids_required"}, 400)
        return

    try:
        id_list = [int(x) for x in ids]
    except Exception:
        handler.send_json({"error": "invalid_ids"}, 400)
        return

    conn = ensure_db()
    rows: List[Dict[str, Any]] = []
    for pid in id_list:
        product = database.get_product(conn, pid)
        if product is not None:
            rows.append(row_to_dict(product))

    if not rows:
        handler.send_json({"error": "products_not_found"}, 404)
        return

    export_path = build_export_filename(EXPORT_DIR)
    wb_data = _build_workbook(rows, export_path)
    filename = export_path.name

    handler.send_response(200)
    handler.send_header(
        "Content-Type",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    handler.send_header("Content-Disposition", f"attachment; filename={filename}")
    handler.send_header("Content-Length", str(len(wb_data)))
    handler.end_headers()
    handler.wfile.write(wb_data)

    duration = time.perf_counter() - start
    logger.info("kalodata minimal export ok products=%d duration=%.3fs", len(rows), duration)


def _build_workbook(
    rows: Sequence[Mapping[str, Any]], output_path: Optional[Path] = None
) -> bytes:
    records: List[Dict[str, Any]] = []
    metadata: List[Dict[str, Any]] = []

    for row in rows:
        record, meta = _convert_row(row)
        records.append(record)
        metadata.append(meta)

    _apply_desire_fallback(records, metadata)

    generated_fields = ("Desire", "Desire Magnitude", "Awareness Level", "Competition Level")
    for record, meta in zip(records, metadata):
        missing: List[str] = []
        for field in generated_fields:
            value = record.get(field)
            if field == "Desire Magnitude":
                if value is None or (isinstance(value, float) and math.isnan(value)):
                    missing.append(field)
            else:
                if not value:
                    missing.append(field)
        if missing:
            logger.warning(
                "Missing generated fields for product %s: %s",
                meta.get("id"),
                ", ".join(missing),
            )

    df = pd.DataFrame(records, columns=BASE_COLUMNS)
    df = reorder_columns(df)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="products", index=False)
        ws = writer.sheets["products"]
        _apply_sheet_format(ws)
        _clear_desire_magnitude_styles(ws)
        _embed_images(ws)

    data = buffer.getvalue()

    if output_path is not None:
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with output_path.open("wb") as fh:
                fh.write(data)
        except Exception:
            logger.exception("failed to persist export workbook to %s", output_path)

    return data


def _apply_sheet_format(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F497D")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    header_map = _build_header_index_map(ws)

    width_preferences = {
        ("image", "img_url", "image url"): 40,
        ("name", "product name"): 40,
        ("category",): 18,
        ("price", "price($)", "price usd", "price$"): 12,
        ("rating", "product rating"): 14,
        ("revenue", "total revenue($)", "total revenue"): 16,
        ("items sold", "item sold", "units sold"): 12,
        ("desire",): 45,
        ("desire magnitude", "desire_mag"): 18,
        ("awareness level",): 18,
        ("competition level",): 18,
        ("tiktokurl", "tiktok url"): 35,
        ("kalodataurl", "kalodata url"): 35,
        ("image preview",): 38,
        ("launch date",): 14,
    }
    for headers, width in width_preferences.items():
        idx = _resolve_column_index(header_map, *headers)
        if idx is not None:
            ws.column_dimensions[get_column_letter(idx)].width = width

    text_column_candidates = [
        ("image", "img_url", "image url"),
        ("name", "product name"),
        ("category",),
        ("desire",),
        ("awareness level",),
        ("competition level",),
    ]
    if ws.max_row >= 2:
        for headers in text_column_candidates:
            col_idx = _resolve_column_index(header_map, *headers)
            if col_idx is None:
                continue
            for row in ws.iter_rows(
                min_row=2,
                max_row=ws.max_row,
                min_col=col_idx,
                max_col=col_idx,
            ):
                cell = row[0]
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        num_fmt_int = "#,##0"
        num_fmt_money = "$#,##0.00"
        num_fmt_rating = "0.0"
        num_fmt_date = "yyyy-mm-dd"
        num_fmt_desire = "0"

        items_col = _resolve_column_index(header_map, "items sold", "item sold", "units sold")
        if items_col is not None:
            for column in ws.iter_cols(items_col, items_col, 2, ws.max_row):
                for cell in column:
                    cell.number_format = num_fmt_int

        price_col = _resolve_column_index(header_map, "price", "price($)", "price usd", "price$")
        if price_col is not None:
            for column in ws.iter_cols(price_col, price_col, 2, ws.max_row):
                for cell in column:
                    cell.number_format = num_fmt_money

        rating_col = _resolve_column_index(header_map, "rating", "product rating")
        if rating_col is not None:
            for column in ws.iter_cols(rating_col, rating_col, 2, ws.max_row):
                for cell in column:
                    cell.number_format = num_fmt_rating

        revenue_col = _resolve_column_index(
            header_map, "revenue", "total revenue($)", "total revenue"
        )
        if revenue_col is not None:
            for column in ws.iter_cols(revenue_col, revenue_col, 2, ws.max_row):
                for cell in column:
                    cell.number_format = num_fmt_money

        launch_col = _resolve_column_index(header_map, "launch date")
        if launch_col is not None:
            for column in ws.iter_cols(launch_col, launch_col, 2, ws.max_row):
                for cell in column:
                    cell.number_format = num_fmt_date

        desire_mag_col = _resolve_column_index(
            header_map, "desire magnitude", "desire_mag"
        )
        if desire_mag_col is not None:
            for column in ws.iter_cols(desire_mag_col, desire_mag_col, 2, ws.max_row):
                for cell in column:
                    cell.number_format = num_fmt_desire

            dv = DataValidation(
                type="whole",
                operator="between",
                formula1="0",
                formula2="100",
                allow_blank=True,
            )
            ws.add_data_validation(dv)
            dv_range = (
                f"{get_column_letter(desire_mag_col)}2:"
                f"{get_column_letter(desire_mag_col)}{ws.max_row}"
            )
            dv.add(dv_range)

    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName="tbl_products", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _cell_range_includes_column(range_string: str, column_index: int) -> bool:
    try:
        cell_range = CellRange(range_string)
    except ValueError:
        return False
    return cell_range.min_col <= column_index <= cell_range.max_col


def _clear_desire_magnitude_styles(ws) -> None:
    if ws.max_row < 2:
        return

    col_idx: Optional[int] = None
    for cell in ws[1]:
        if _is_desire_magnitude_header(cell.value):
            col_idx = cell.col_idx
            break

    if col_idx is None:
        header_map = _build_header_index_map(ws)
        col_idx = _resolve_column_index(
            header_map,
            "desire magnitude",
            "desire_mag",
            "desire mag",
            "magnitud del deseo",
        )

    if col_idx is None:
        return

    try:
        cf = ws.conditional_formatting
        cf_rules = getattr(cf, "_cf_rules", None)
        if cf_rules:
            keep: List[Tuple[str, Any]] = []
            for key, rules in list(cf_rules.items()):
                ranges = str(key).replace(",", " ").split()
                if any(_cell_range_includes_column(rng, col_idx) for rng in ranges):
                    continue
                keep.append((key, rules))
            cf_rules.clear()
            for key, rules in keep:
                cf_rules[key] = rules
    except Exception:
        logger.exception("failed to prune conditional formats for desire magnitude")

    for row in ws.iter_rows(
        min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row
    ):
        cell = row[0]
        cell.fill = PatternFill()


def _convert_row(row: Mapping[str, Any]) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    extras = _ensure_dict(row.get("extra"))
    sources = _prepare_sources(row, extras)

    product_name = _coerce_text(_value_from_sources(sources, TEXT_FIELD_KEYS["Product Name"]))
    product_name = sanitize_product_name(product_name) or ""
    tiktok_url = _coerce_text(_value_from_sources(sources, TEXT_FIELD_KEYS["TikTokUrl"]))
    kalodata_url = _coerce_text(_value_from_sources(sources, TEXT_FIELD_KEYS["KalodataUrl"]))
    image_url = _coerce_text(_value_from_sources(sources, TEXT_FIELD_KEYS["Img_url"]))
    category = _coerce_text(_value_from_sources(sources, TEXT_FIELD_KEYS["Category"]))

    price_val = _parse_number(_value_from_sources(sources, NUMERIC_FIELD_KEYS["Price($)"]))
    price = round(price_val, 2) if price_val is not None else None

    rating_val = _parse_number(_value_from_sources(sources, NUMERIC_FIELD_KEYS["Product Rating"]))
    rating = round(rating_val, 2) if rating_val is not None else None

    units_val = _parse_number(_value_from_sources(sources, NUMERIC_FIELD_KEYS["Item Sold"]))
    units = int(round(units_val)) if units_val is not None else None

    total_revenue = _compute_total_revenue(sources)

    launch_date = _normalize_launch_date(
        _value_from_sources(sources, TEXT_FIELD_KEYS["Launch Date"])
    )

    desire_text_raw = _value_from_sources(sources, DESIRE_TEXT_KEYS)
    desire_text = _coerce_text(desire_text_raw)

    desire_mag_raw = _value_from_sources(sources, DESIRE_MAGNITUDE_KEYS)
    desire_score: Optional[int]
    if desire_mag_raw is not None:
        desire_score = _normalize_desire_score(desire_mag_raw)
    else:
        desire_score_raw = _value_from_sources(sources, DESIRE_SCORE_KEYS)
        if desire_score_raw is not None:
            desire_score = _normalize_desire_score(desire_score_raw)
        elif _looks_numeric(desire_text_raw):
            desire_score = _normalize_desire_score(desire_text_raw)
        else:
            desire_score = None

    awareness_raw = _value_from_sources(sources, AWARENESS_KEYS)
    awareness = _normalize_awareness(awareness_raw)

    competition_raw = _value_from_sources(sources, COMPETITION_KEYS)
    competition = _normalize_competition(competition_raw)

    record: Dict[str, Any] = {
        "Product Name": product_name,
        "TikTokUrl": tiktok_url,
        "KalodataUrl": kalodata_url,
        "Img_url": image_url,
        "Image Preview": None,
        "Category": category,
        "Price($)": price,
        "Product Rating": rating,
        "Item Sold": units,
        "Total Revenue($)": total_revenue,
        "Launch Date": launch_date,
        "Desire": desire_text,
        "Desire Magnitude": desire_score,
        "Awareness Level": awareness,
        "Competition Level": competition,
    }

    meta: Dict[str, Any] = {
        "id": row.get("id"),
        "item_sold": units,
        "rating": rating,
    }

    return record, meta


def _apply_desire_fallback(records: Sequence[Dict[str, Any]], metadata: Sequence[Dict[str, Any]]) -> None:
    items = sorted(v for v in (meta.get("item_sold") for meta in metadata) if v is not None)
    ratings = sorted(v for v in (meta.get("rating") for meta in metadata) if v is not None)

    for record, meta in zip(records, metadata):
        magnitude = record.get("Desire Magnitude")
        if magnitude is None or (isinstance(magnitude, float) and math.isnan(magnitude)):
            item_pct = _rank_percentile(meta.get("item_sold"), items)
            rating_pct = _rank_percentile(meta.get("rating"), ratings)
            fallback = int(round(((item_pct + rating_pct) / 2.0) * 100))
            record["Desire Magnitude"] = max(0, min(100, fallback))
        else:
            record["Desire Magnitude"] = int(round(float(magnitude)))


def _rank_percentile(value: Any, sorted_values: Sequence[float]) -> float:
    if value is None or not sorted_values:
        return 0.5
    n = len(sorted_values)
    if n == 1:
        return 1.0
    insert = bisect.bisect_left(sorted_values, value)
    if insert >= n:
        return 1.0
    hi = bisect.bisect_right(sorted_values, value) - 1
    if hi >= insert:
        avg_rank = (insert + hi) / 2.0
    else:
        if insert == 0:
            return 0.0
        lower_index = insert - 1
        upper_index = insert
        lower_val = sorted_values[lower_index]
        upper_val = sorted_values[upper_index]
        if upper_val == lower_val:
            avg_rank = float(lower_index)
        else:
            fraction = (float(value) - lower_val) / (upper_val - lower_val)
            avg_rank = lower_index + fraction
    percentile = avg_rank / (n - 1)
    return max(0.0, min(1.0, float(percentile)))


def _embed_images(ws) -> None:
    header_map = _build_header_index_map(ws)
    image_col = _resolve_column_index(header_map, "image", "img_url", "image url")
    preview_col = _resolve_column_index(header_map, "image preview")

    if image_col is None or preview_col is None:
        logger.warning("Unable to locate image columns for embedding")
        return

    column_letter = get_column_letter(preview_col)
    ws.column_dimensions[column_letter].width = 38
    for row_idx in range(2, ws.max_row + 1):
        url = ws.cell(row=row_idx, column=image_col).value
        bio = _fetch_and_resize(url)
        target_cell = ws.cell(row=row_idx, column=preview_col)
        if bio is not None:
            pic = XLImage(bio)
            anchor = f"{column_letter}{row_idx}"
            ws.add_image(pic, anchor)
            ws.row_dimensions[row_idx].height = _IMG_ROW_HEIGHT
            target_cell.value = None
            target_cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            target_cell.value = "(no image)"
            target_cell.font = Font(color="FF888888", italic=True)
            target_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[row_idx].height = 45


def _fetch_and_resize(url: Any) -> Optional[io.BytesIO]:
    if not url:
        return None
    url_str = str(url)
    if url_str in _IMG_CACHE:
        cached = _IMG_CACHE[url_str]
        return io.BytesIO(cached) if cached is not None else None
    try:
        with _IMAGE_SESSION.get(url_str, timeout=5, stream=True) as response:
            response.raise_for_status()
            raw_bytes = response.content
        img = PILImage.open(io.BytesIO(raw_bytes)).convert("RGBA")
        img.thumbnail((_IMG_MAX_WIDTH, _IMG_MAX_HEIGHT))
        output = io.BytesIO()
        img.save(output, format="PNG")
        data = output.getvalue()
        _IMG_CACHE[url_str] = data
        return io.BytesIO(data)
    except Exception:
        _IMG_CACHE[url_str] = None
        return None


def _looks_numeric(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return False
        return _parse_number(text) is not None
    return False

def _prepare_sources(*dicts: Mapping[str, Any]) -> List[Dict[str, Any]]:
    sources: List[Dict[str, Any]] = []
    for data in dicts:
        mapping = _ensure_dict(data)
        if not mapping:
            continue
        direct: Dict[str, Any] = {}
        lower: Dict[str, Any] = {}
        sanitized: Dict[str, Any] = {}
        for key, value in mapping.items():
            if not isinstance(key, str):
                continue
            direct[key] = value
            lower[key.lower()] = value
            sanitized[_sanitize_key(key)] = value
        sources.extend([direct, lower, sanitized])
    return sources


def _value_from_sources(sources: Sequence[Mapping[str, Any]], keys: Iterable[str]) -> Any:
    for key in keys:
        for variant in _key_variants(key):
            for source in sources:
                if variant in source:
                    value = source[variant]
                    if not _is_missing(value):
                        return value
    return None


def _normalize_desire_score(raw: Any) -> Optional[int]:
    num = _parse_number(raw)
    if num is None:
        return None
    if num <= 1.0:
        num *= 100.0
    num = max(0.0, min(100.0, num))
    return int(round(num))


def _normalize_awareness(raw: Any) -> str:
    if raw is None:
        return ""
    if isinstance(raw, str):
        stripped = raw.strip()
        if not stripped:
            return ""
        parsed = _parse_number(stripped)
        if parsed is not None and not math.isnan(parsed):
            return _awareness_from_numeric(parsed)
        return stripped
    num = _parse_number(raw)
    if num is None or math.isnan(num):
        return ""
    return _awareness_from_numeric(num)


def _awareness_from_numeric(value: float) -> str:
    if value <= 1.0:
        scaled = int(round(value * 4))
    else:
        scaled = int(round(value))
    scaled = max(0, min(4, scaled))
    return _AWARENESS_LABELS.get(scaled, "")


def _normalize_competition(raw: Any) -> str:
    if raw is None:
        return ""
    if isinstance(raw, str):
        stripped = raw.strip()
        if not stripped:
            return ""
        parsed = _parse_number(stripped)
        if parsed is None or math.isnan(parsed):
            return stripped
        return _map_competition(parsed)
    num = _parse_number(raw)
    if num is None or math.isnan(num):
        return ""
    return _map_competition(num)


def _map_competition(value: float) -> str:
    if value <= 1.0:
        score = value
    elif value <= 100.0:
        score = value / 100.0
    else:
        score = value / 100.0
    if score < 0.35:
        return "Low"
    if score < 0.7:
        return "Medium"
    return "High"


def _compute_total_revenue(sources: Sequence[Mapping[str, Any]]) -> float:
    total = 0.0
    for key in (
        "Revenue($)",
        "Live Revenue($)",
        "Video Revenue($)",
        "Shopping Mall Revenue($)",
    ):
        raw = _value_from_sources(sources, NUMERIC_FIELD_KEYS[key])
        total += _parse_money(raw)
    return round(total, 2)


def _normalize_launch_date(raw: Any) -> Any:
    if raw is None:
        return ""
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, (int, float)):
        if isinstance(raw, float) and math.isnan(raw):
            return ""
        try:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=float(raw))).date()
        except Exception:
            return str(raw)
    text = str(raw).strip()
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text).date()
    except Exception:
        pass
    candidate = text.split()[0]
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(candidate, fmt).date()
        except Exception:
            continue
    return text


def _parse_money(value: Any) -> float:
    num = _parse_number(value)
    if num is None:
        return 0.0
    return float(num)

def _parse_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    multiplier = 1.0
    last = text[-1]
    if last in {"k", "K", "m", "M"}:
        multiplier = 1_000.0 if last in {"k", "K"} else 1_000_000.0
        text = text[:-1]
    text = _NUMERIC_STRIP_RE.sub("", text)
    if not text:
        return None
    try:
        return float(text) * multiplier
    except Exception:
        return None


def _coerce_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value)


def _key_variants(key: str) -> List[str]:
    variants = [key]
    lower = key.lower()
    if lower not in variants:
        variants.append(lower)
    sanitized = _sanitize_key(key)
    if sanitized not in variants:
        variants.append(sanitized)
    compact = key.replace(" ", "")
    if compact not in variants:
        variants.append(compact)
    compact_lower = compact.lower()
    if compact_lower not in variants:
        variants.append(compact_lower)
    return variants


def _sanitize_key(name: str) -> str:
    return "".join(ch.lower() for ch in name if ch.isalnum())


def _ensure_dict(value: Any) -> Dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except Exception:
            return {}
        return parsed if isinstance(parsed, dict) else {}
    if isinstance(value, Mapping):
        return dict(value)
    return {}


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return True
        return stripped.lower() in {"na", "n/a", "none", "null"}
    return False


def _read_length(headers: MutableMapping[str, str]) -> int:
    try:
        return int(headers.get("Content-Length", "0"))
    except Exception:
        return 0
